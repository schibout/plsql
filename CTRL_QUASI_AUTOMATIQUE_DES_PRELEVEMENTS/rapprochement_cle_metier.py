#!/usr/bin/env python3
"""Rapprochement des prelevements Oracle / EDF par cle metier.

Remplace le rapprochement par agregat journalier (qui projetait la date Oracle
vers une date EDF via une heuristique J+2/J+4) par un rapprochement sur la cle
metier exacte portee par les deux sources :

    (IBAN creancier, date d'echeance)

    Oracle : ENTITYBANKACCOUNTNUMBER x TRANSACTIONDATE
    EDF    : IBAN CREANCIER          x DATE D'ECHEANCE

Les rejets internes sont apparies ligne a ligne sur le RUM (SEPAMANDATEID), ce
qui permet de qualifier chaque ecart au lieu de le constater.

Usage : python rapprochement_cle_metier.py [--date AAAA-MM-JJ] [--jours 10]

Codes retour :
    0 = rien a traiter
    1 = anomalies detectees
    2 = erreur de traitement (dont lignes Oracle non conformes)
    3 = execution degradee (resultat partiel, a ne pas confondre avec un run propre)
"""
import argparse
import csv
import fnmatch
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:  # pragma: no cover
    sys.exit("Le module 'openpyxl' est requis. Installez-le avec :\n"
             "    python -m pip install -r requirements.txt")

# Helpers deja valides par l'outil existant : lecture, mise en forme, couleurs.
from prelevements_rapprochement import (
    lire_lignes, _titre, _entetes, _ajuster_largeurs, _colonnes_texte,
    COULEUR_ENTETE, COULEUR_ECART, COULEUR_OK, FORMAT_MONTANT, FORMAT_NOMBRE,
    FORMAT_TEXTE, TAILLE_BASE,
)

# --- Format Oracle ---------------------------------------------------------
NB_COLONNES_ORACLE = 55
COLONNES_REQUISES = ("TRANSACTIONDATE", "AMOUNT", "ENTITYBANKACCOUNTNUMBER",
                     "SEPAMANDATEID", "COUNTERPARTYBANKACCOUNTNUMBER",
                     "COUNTERPARTYNAME")
RE_IBAN_FR = re.compile(r"^FR\d{12}")
# Date d'emission portee par le nom de fichier : DK_..-<AAAAMMJJ>-<id>_<AAAAMMJJ>-<heure>
RE_DATE_FICHIER = re.compile(r"-(\d{8})-\d+_")

# Delai emission -> echeance observe : 1 a 27 jours. La borne avant du perimetre
# doit couvrir cet horizon, sinon des cles emises sont exclues du rapport.
MARGE_ECHEANCE_FUTURE = 25
# Une confirmation EDF arrive toujours dans le 1er ou le 2e fichier posterieur a
# l'emission. Au-dela, l'absence n'est plus un delai mais une anomalie.
SEUIL_FICHIERS_EDF = 2
# Delai maximal constate entre l'emission Oracle et le fichier EDF qui la
# confirme (5 jours calendaires sur 264 lots). Sert a savoir si l'emission
# correspondant a une remontee EDF pouvait se trouver dans l'historique.
LAG_MAX_EDF = 5

COULEUR_ATTENTE = "FFFFF2CC"   # jaune pale
COULEUR_SIGNALE = "FFFCE4D6"   # peche

# --- Statuts ---------------------------------------------------------------
RAPPROCHE = "RAPPROCHE"
RAPPROCHE_REJET_POSTERIEUR = "RAPPROCHE_AVEC_REJET_POSTERIEUR"
EXPLIQUE_PAR_REJET = "EXPLIQUE_PAR_REJET"
REJETE_INTEGRALEMENT = "REJETE_INTEGRALEMENT"
REJET_PARTIEL_NON_CONFIRME = "REJET_PARTIEL_NON_CONFIRME"
EN_ATTENTE = "EN_ATTENTE"
NON_RECU = "NON_RECU"
ECART_PARTIEL = "ECART_PARTIEL"
EDF_SANS_ORACLE = "EDF_SANS_ORACLE"
HORS_PERIMETRE_HISTORIQUE = "HORS_PERIMETRE_HISTORIQUE"

# Seuls ces statuts appellent une action.
STATUTS_ANOMALIE = {NON_RECU, ECART_PARTIEL, EDF_SANS_ORACLE}
# Conformes arithmetiquement, mais le metier doit les voir.
STATUTS_SIGNALES = {RAPPROCHE_REJET_POSTERIEUR, REJET_PARTIEL_NON_CONFIRME}

ORDRE_STATUTS = [
    RAPPROCHE, RAPPROCHE_REJET_POSTERIEUR, EXPLIQUE_PAR_REJET,
    REJETE_INTEGRALEMENT, REJET_PARTIEL_NON_CONFIRME, EN_ATTENTE,
    NON_RECU, ECART_PARTIEL, EDF_SANS_ORACLE, HORS_PERIMETRE_HISTORIQUE,
]

EXPLICATIONS = {
    RAPPROCHE: "Nombre et montant identiques de part et d'autre.",
    RAPPROCHE_REJET_POSTERIEUR: "Totaux conformes, mais un rejet est arrive apres la remontee EDF : le prelevement echouera.",
    EXPLIQUE_PAR_REJET: "L'ecart correspond exactement aux rejets internes.",
    REJETE_INTEGRALEMENT: "Tous les prelevements ont ete rejetes : EDF ne remonte donc aucune ligne.",
    REJET_PARTIEL_NON_CONFIRME: "Des rejets existent mais EDF n'a encore rien remonte pour cette cle.",
    EN_ATTENTE: "Emis, pas encore confirme par EDF, dans le delai normal.",
    NON_RECU: "Emis, non confirme par EDF au-dela du delai normal.",
    ECART_PARTIEL: "Ecart non explique par les rejets.",
    EDF_SANS_ORACLE: "EDF a remonte des prelevements sans contrepartie Oracle.",
    HORS_PERIMETRE_HISTORIQUE: "Echeance anterieure a l'historique Oracle disponible : non concluant.",
}


class ErreurTraitement(Exception):
    """Erreur fonctionnelle : interrompt le traitement avec un message clair."""


@dataclass
class LigneOracle:
    iban_creancier: str
    echeance: date
    montant: Decimal
    rum: str
    iban_debiteur: str
    nom: str
    emission: date
    fichier: str


@dataclass
class LigneRejet:
    iban_creancier: str
    rum: str
    iban_debiteur: str
    echeance: date
    montant: Decimal
    code: str
    motif: str
    fichier: str
    date_fichier: date
    appariee: bool = False


@dataclass
class Tranche:
    date_fichier: date
    nb: int
    montant: Decimal


@dataclass
class AgregatEdf:
    nb: int = 0
    montant: Decimal = Decimal(0)
    tranches: list = field(default_factory=list)


class Diagnostic:
    """Compteurs de traitement : un rapport partiel doit se voir."""

    def __init__(self):
        self.fichiers_oracle = 0
        self.lignes_oracle = 0
        self.fichiers_edf = 0
        self.fichiers_edf_sans_si = 0
        self.lignes_edf = 0
        self.fichiers_rejets = 0
        self.lignes_rejets = 0
        self.rejets_dupliques = 0
        self.avertissements = []

    def avertir(self, message):
        self.avertissements.append(message)
        print(f"AVERTISSEMENT : {message}", file=sys.stderr)


# ---------------------------------------------------------------------------
# Analyse : conversions strictes
# ---------------------------------------------------------------------------
def date_us(valeur):
    """Oracle : MM/JJ/AAAA. Leve si le format ne colle pas."""
    return datetime.strptime(valeur.strip(), "%m/%d/%Y").date()


def date_fr(valeur):
    """EDF et rejets : JJ/MM/AAAA."""
    return datetime.strptime(valeur.strip(), "%d/%m/%Y").date()


def date_compacte(valeur):
    """Noms de dossiers et de fichiers : AAAAMMJJ."""
    return datetime.strptime(valeur.strip(), "%Y%m%d").date()


def montant_oracle(valeur):
    """Oracle : separateur decimal point, aucun separateur de milliers."""
    v = valeur.strip()
    if "," in v:
        raise InvalidOperation(f"virgule inattendue dans un montant Oracle : {valeur!r}")
    return Decimal(v)


def montant_edf(valeur):
    """EDF et rejets : separateur decimal virgule, aucun separateur de milliers."""
    v = valeur.strip()
    if "." in v:
        raise InvalidOperation(f"point inattendu dans un montant EDF : {valeur!r}")
    return Decimal(v.replace(",", "."))


def colonnes_oracle(lignes, fichier):
    """Resout les index depuis l'en-tete FORMAT1ENTITYID.

    Jamais d'index code en dur : un changement de format Oracle doit arreter le
    traitement, pas produire un rapport faux.
    """
    for ligne in lignes:
        if ligne.startswith("FORMAT1ENTITYID"):
            noms = [c.strip() for c in ligne.split(",")]
            manquantes = [c for c in COLONNES_REQUISES if c not in noms]
            if manquantes:
                raise ErreurTraitement(
                    f"Colonnes absentes de l'en-tete de {fichier} : {', '.join(manquantes)}")
            return {c: noms.index(c) for c in COLONNES_REQUISES}
    raise ErreurTraitement(f"Ligne d'en-tete FORMAT1ENTITYID absente de {fichier}")


def date_emission(fichier, diag):
    """La date d'emission est portee par le NOM du fichier, pas par le dossier.

    Le dossier ORACLE/<date> est systematiquement a J-1, et certains dossiers
    melangent deux dates d'emission (ex. ORACLE/20260626).
    """
    m = RE_DATE_FICHIER.search(fichier.name)
    if m:
        try:
            return date_compacte(m.group(1))
        except ValueError:
            pass
    try:
        repli = date_compacte(fichier.parent.name)
    except ValueError:
        diag.avertir(f"Date d'emission introuvable pour {fichier.name}")
        return None
    diag.avertir(f"Date d'emission deduite du dossier pour {fichier.name} : {repli}")
    return repli


# ---------------------------------------------------------------------------
# Chargement
# ---------------------------------------------------------------------------
def charger_oracle(oracle_path, motifs, diag):
    """Retourne (lignes valides, lignes non conformes)."""
    if not oracle_path.is_dir():
        raise ErreurTraitement(f"Dossier Oracle introuvable : {oracle_path}")

    lignes_ok, lignes_ko = [], []
    vu_jour_sup_12 = False

    for fichier in sorted(oracle_path.rglob("*")):
        if not fichier.is_file():
            continue
        if not any(fnmatch.fnmatch(fichier.name, m) for m in motifs):
            continue

        contenu = lire_lignes(fichier)
        cols = colonnes_oracle(contenu, fichier.name)
        emission = date_emission(fichier, diag)
        if emission is None:
            continue
        diag.fichiers_oracle += 1

        for num, ligne in enumerate(contenu, start=1):
            if not ligne.strip():
                continue
            if ligne.startswith("HEADER") or ligne.startswith("FORMAT1ENTITYID"):
                continue

            champs = ligne.split(",")
            # ENTITYBANKACCOUNTNUMBER est en position 11, apres deux champs de
            # texte libre : une virgule saisie dans une reference decalerait
            # toute la cle. On valide plutot que de faire confiance.
            if len(champs) != NB_COLONNES_ORACLE:
                lignes_ko.append({
                    "fichier": fichier.name, "ligne": num,
                    "motif": f"{len(champs)} champs au lieu de {NB_COLONNES_ORACLE}",
                    "extrait": ligne[:200]})
                continue

            iban = champs[cols["ENTITYBANKACCOUNTNUMBER"]].strip()
            if not RE_IBAN_FR.match(iban):
                lignes_ko.append({
                    "fichier": fichier.name, "ligne": num,
                    "motif": f"IBAN creancier invalide : {iban!r}",
                    "extrait": ligne[:200]})
                continue

            brut_date = champs[cols["TRANSACTIONDATE"]].strip()
            try:
                echeance = date_us(brut_date)
                montant = montant_oracle(champs[cols["AMOUNT"]])
            except (ValueError, InvalidOperation) as exc:
                lignes_ko.append({
                    "fichier": fichier.name, "ligne": num,
                    "motif": f"date ou montant illisible : {exc}",
                    "extrait": ligne[:200]})
                continue

            # Preuve du format MM/JJ/AAAA : le 2e composant (le jour) doit
            # depasser 12 au moins une fois. Tant que ce n'est pas le cas, les
            # deux conventions restent indiscernables.
            if int(brut_date.split("/")[1]) > 12:
                vu_jour_sup_12 = True

            lignes_ok.append(LigneOracle(
                iban_creancier=iban,
                echeance=echeance,
                montant=montant,
                rum=champs[cols["SEPAMANDATEID"]].strip(),
                iban_debiteur=champs[cols["COUNTERPARTYBANKACCOUNTNUMBER"]].strip(),
                nom=champs[cols["COUNTERPARTYNAME"]].strip(),
                emission=emission,
                fichier=fichier.name))
            diag.lignes_oracle += 1

    if not lignes_ok and not lignes_ko:
        raise ErreurTraitement(f"Aucune ligne Oracle exploitable sous {oracle_path}.")

    # Oracle est cense etre en MM/JJ/AAAA. Si aucun mois > 12 n'apparait, le
    # format a peut-etre bascule en JJ/MM : toutes les echeances seraient fausses
    # sans qu'aucune exception ne soit levee.
    if lignes_ok and not vu_jour_sup_12:
        diag.avertir("Aucune TRANSACTIONDATE avec un premier composant > 12 : "
                     "le format Oracle MM/JJ/AAAA n'est pas confirme, les echeances "
                     "sont peut-etre inversees.")

    return lignes_ok, lignes_ko


def charger_edf(edf_path, motif, nom_si, diag):
    """Retourne (agregats par cle, dates des fichiers EDF lus).

    Une cle presente dans plusieurs fichiers EDF correspond a des remontees
    INCREMENTALES : elles se somment (verifie : 102 + 1 = 103 cote Oracle).
    """
    agregats = defaultdict(AgregatEdf)
    dates_fichiers = []

    if not edf_path.is_dir():
        diag.avertir(f"Dossier EDF introuvable : {edf_path}.")
        return agregats, dates_fichiers

    for fichier in sorted(edf_path.glob(motif)):
        if not fichier.is_file():
            continue
        segments = fichier.name.split(".")
        try:
            date_fichier = date_compacte(segments[1])
        except (IndexError, ValueError):
            diag.avertir(f"Date illisible dans le nom du fichier EDF, ignore : {fichier.name}")
            continue

        dates_fichiers.append(date_fichier)
        diag.fichiers_edf += 1
        lignes_si = 0

        for ligne in lire_lignes(fichier):
            champs = ligne.split(";")
            if len(champs) < 5 or champs[0].strip() != nom_si:
                continue
            try:
                echeance = date_fr(champs[2])
                nb = int(champs[3].strip())
                montant = montant_edf(champs[4])
            except (ValueError, InvalidOperation) as exc:
                diag.avertir(f"Ligne EDF illisible dans {fichier.name} ({exc}) : {ligne}")
                continue

            cle = (champs[1].strip(), echeance)
            agg = agregats[cle]
            agg.nb += nb
            agg.montant += montant
            agg.tranches.append(Tranche(date_fichier, nb, montant))
            diag.lignes_edf += 1
            lignes_si += 1

        # Un fichier present sans ligne du SI est NORMAL (journee sans
        # prelevement pour ce SI) : on le compte sans degrader l'execution.
        # C'est l'ABSENCE de fichier qui signale un incident de flux.
        if lignes_si == 0:
            diag.fichiers_edf_sans_si += 1

    return agregats, sorted(dates_fichiers)


def charger_rejets(rejets_path, motif, diag):
    """Charge les rejets en dedoublonnant les republications a l'identique.

    Un meme rejet peut etre republie dans un fichier ulterieur (verifie). Deux
    RUM identiques sur des echeances DIFFERENTES sont en revanche legitimes :
    l'echeance fait donc partie de la signature.
    """
    rejets = []
    if not rejets_path.is_dir():
        diag.avertir(f"Dossier de rejets introuvable : {rejets_path}.")
        return rejets

    vues = set()
    fichiers = sorted(p for p in rejets_path.iterdir()
                      if p.is_file() and fnmatch.fnmatch(p.name, motif))
    diag.fichiers_rejets = len(fichiers)

    for fichier in fichiers:
        segments = fichier.name.split(".")
        try:
            date_fichier = date_compacte(segments[1])
        except (IndexError, ValueError):
            diag.avertir(f"Date illisible dans le nom du fichier de rejets : {fichier.name}")
            continue

        for ligne in lire_lignes(fichier):
            champs = ligne.split(";")
            if len(champs) < 7 or not champs[0].strip().startswith("FR"):
                continue
            try:
                echeance = date_fr(champs[3])
                montant = montant_edf(champs[4])
            except (ValueError, InvalidOperation) as exc:
                diag.avertir(f"Ligne de rejet illisible dans {fichier.name} ({exc}) : {ligne}")
                continue

            signature = (champs[0].strip(), champs[1].strip(), champs[2].strip(),
                         echeance, montant)
            if signature in vues:
                diag.rejets_dupliques += 1
                continue
            vues.add(signature)

            rejets.append(LigneRejet(
                iban_creancier=champs[0].strip(), rum=champs[1].strip(),
                iban_debiteur=champs[2].strip(), echeance=echeance, montant=montant,
                code=champs[5].strip(), motif=champs[6].strip(),
                fichier=fichier.name, date_fichier=date_fichier))
            diag.lignes_rejets += 1

    return rejets


# ---------------------------------------------------------------------------
# Appariement des rejets sur le RUM
# ---------------------------------------------------------------------------
def apparier_rejets(lignes_oracle, rejets):
    """Apparie chaque rejet a une ligne Oracle sur (IBAN creancier, RUM, echeance).

    Indispensable : les fichiers de rejets ne portent pas le nom du SI, et
    plusieurs IBAN creanciers sont partages entre CIF et ORACLE. Le RUM est le
    seul discriminant fiable ; un rejet non apparie releve d'un autre SI.
    """
    index = defaultdict(list)
    for lo in lignes_oracle:
        index[(lo.iban_creancier, lo.rum, lo.echeance)].append(lo)

    apparies = defaultdict(list)   # cle (iban, echeance) -> [LigneRejet]
    for rejet in rejets:
        candidats = index.get((rejet.iban_creancier, rejet.rum, rejet.echeance))
        if not candidats:
            continue
        rejet.appariee = True
        apparies[(rejet.iban_creancier, rejet.echeance)].append(rejet)
    return apparies


# ---------------------------------------------------------------------------
# Machine a etats
# ---------------------------------------------------------------------------
def nb_fichiers_edf_depuis(emission, dates_fichiers_edf, reference):
    return sum(1 for d in dates_fichiers_edf if emission < d <= reference)


def statut_cle(ora_nb, ora_montant, edf, rejets_cle, derniere_emission,
               dates_edf, reference, couverture_debut, echeance):
    """Retourne (statut, ecart_nb, ecart_montant). Ordre d'evaluation significatif."""
    rej_nb = len(rejets_cle)
    rej_montant = sum((r.montant for r in rejets_cle), Decimal(0))

    if edf is None:
        # EDF muet : distinguer rejet integral, rejet partiel, attente et anomalie.
        if rej_nb and (rej_nb, rej_montant) == (ora_nb, ora_montant):
            return REJETE_INTEGRALEMENT, -ora_nb, -ora_montant
        if rej_nb:
            return REJET_PARTIEL_NON_CONFIRME, -ora_nb, -ora_montant
        if derniere_emission is not None and \
                nb_fichiers_edf_depuis(derniere_emission, dates_edf, reference) > SEUIL_FICHIERS_EDF:
            return NON_RECU, -ora_nb, -ora_montant
        return EN_ATTENTE, -ora_nb, -ora_montant

    ecart_nb = edf.nb - ora_nb
    ecart_montant = edf.montant - ora_montant

    if ora_nb == 0:
        # EDF a remonte sans contrepartie Oracle. Ce n'est concluant que si
        # l'emission correspondante tombait dans l'historique disponible : une
        # remontee EDF datee du premier jour de l'archive confirme forcement une
        # emission anterieure a celle-ci, qu'on ne peut ni trouver ni incriminer.
        if couverture_debut is not None and edf.tranches:
            premiere_remontee = min(t.date_fichier for t in edf.tranches)
            if premiere_remontee - timedelta(days=LAG_MAX_EDF) < couverture_debut:
                return HORS_PERIMETRE_HISTORIQUE, ecart_nb, ecart_montant
        return EDF_SANS_ORACLE, ecart_nb, ecart_montant

    if (ecart_nb, ecart_montant) == (0, Decimal(0)):
        # Conforme, mais un rejet arrive apres la remontee EDF n'y est pas reflete.
        return (RAPPROCHE_REJET_POSTERIEUR if rej_nb else RAPPROCHE), 0, Decimal(0)

    if rej_nb and (-ecart_nb, -ecart_montant) == (rej_nb, rej_montant):
        return EXPLIQUE_PAR_REJET, ecart_nb, ecart_montant

    return ECART_PARTIEL, ecart_nb, ecart_montant


def construire_rapprochement(lignes_oracle, agregats_edf, rejets_par_cle,
                             dates_edf, reference, jours):
    """Une ligne par cle (IBAN creancier, echeance), dans le perimetre demande."""
    debut = reference - timedelta(days=jours)
    fin = reference + timedelta(days=MARGE_ECHEANCE_FUTURE)

    ora = defaultdict(lambda: {"nb": 0, "montant": Decimal(0), "emissions": set()})
    for lo in lignes_oracle:
        e = ora[(lo.iban_creancier, lo.echeance)]
        e["nb"] += 1
        e["montant"] += lo.montant
        e["emissions"].add(lo.emission)

    couverture_debut = min((lo.emission for lo in lignes_oracle), default=None)

    resultat = []
    for cle in sorted(set(ora) | set(agregats_edf), key=lambda k: (k[1], k[0])):
        iban, echeance = cle
        if not (debut <= echeance <= fin):
            continue

        o = ora.get(cle)
        ora_nb = o["nb"] if o else 0
        ora_montant = o["montant"] if o else Decimal(0)
        emissions = sorted(o["emissions"]) if o else []
        edf = agregats_edf.get(cle)
        rejets_cle = rejets_par_cle.get(cle, [])

        statut, ecart_nb, ecart_montant = statut_cle(
            ora_nb, ora_montant, edf, rejets_cle,
            emissions[-1] if emissions else None,
            dates_edf, reference, couverture_debut, echeance)

        resultat.append({
            "iban_creancier": iban,
            "echeance": echeance,
            "nb_oracle": ora_nb,
            "montant_oracle": ora_montant,
            "nb_edf": edf.nb if edf else 0,
            "montant_edf": edf.montant if edf else Decimal(0),
            "ecart_nb": ecart_nb,
            "ecart_montant": ecart_montant,
            "nb_rejets": len(rejets_cle),
            "montant_rejets": sum((r.montant for r in rejets_cle), Decimal(0)),
            "codes_rejets": " ".join(sorted({r.code for r in rejets_cle})),
            "statut": statut,
            "emissions": " + ".join(d.strftime("%d/%m/%Y") for d in emissions),
            "tranches_edf": " + ".join(
                f"{t.date_fichier.strftime('%d/%m')}:{t.nb}" for t in edf.tranches) if edf else "",
        })
    return resultat


# ---------------------------------------------------------------------------
# Sorties
# ---------------------------------------------------------------------------
def _f(valeur):
    return float(round(valeur, 2))


def generer_classeur(chemin, rapprochement, rejets, lignes_ko, resume, contexte):
    wb = Workbook()

    # --- Synthese ---
    ws = wb.active
    ws.title = "Synthèse"
    _titre(ws, "A1", "RAPPROCHEMENT DES PRÉLÈVEMENTS PAR CLÉ MÉTIER")
    ligne = 3
    for libelle, valeur in contexte:
        ws.cell(row=ligne, column=1, value=libelle).font = Font(bold=True, size=TAILLE_BASE)
        ws.cell(row=ligne, column=2, value=valeur)
        ligne += 1

    ligne += 1
    _entetes(ws, ligne, 1, ["Statut", "Nb clés", "Nb prélèvements",
                            "Montant (€)", "Signification"], COULEUR_ENTETE)
    depart = ligne + 1
    for statut in ORDRE_STATUTS:
        infos = resume.get(statut)
        if not infos:
            continue
        ws.cell(row=depart, column=1, value=statut)
        ws.cell(row=depart, column=2, value=infos["cles"]).number_format = FORMAT_NOMBRE
        ws.cell(row=depart, column=3, value=infos["nb"]).number_format = FORMAT_NOMBRE
        ws.cell(row=depart, column=4, value=_f(infos["montant"])).number_format = FORMAT_MONTANT
        ws.cell(row=depart, column=5, value=EXPLICATIONS[statut])
        if statut in STATUTS_ANOMALIE:
            for col in range(1, 6):
                ws.cell(row=depart, column=col).fill = PatternFill("solid", fgColor=COULEUR_ECART)
            ws.cell(row=depart, column=1).font = Font(bold=True, size=TAILLE_BASE)
        elif statut in STATUTS_SIGNALES:
            ws.cell(row=depart, column=1).fill = PatternFill("solid", fgColor=COULEUR_SIGNALE)
        elif statut == EN_ATTENTE:
            ws.cell(row=depart, column=1).fill = PatternFill("solid", fgColor=COULEUR_ATTENTE)
        else:
            ws.cell(row=depart, column=1).fill = PatternFill("solid", fgColor=COULEUR_OK)
        depart += 1
    _ajuster_largeurs(ws, {"A": 34, "E": 80}, maxi=80)

    # --- Rapprochement ---
    ws2 = wb.create_sheet("Rapprochement")
    _titre(ws2, "A1", "DÉTAIL PAR CLÉ (IBAN CRÉANCIER × ÉCHÉANCE)")
    _entetes(ws2, 3, 1, [
        "IBAN Créancier", "Échéance", "Nb Oracle", "Montant Oracle (€)",
        "Nb EDF", "Montant EDF (€)", "Écart Nb", "Écart Montant (€)",
        "Nb Rejets", "Montant Rejets (€)", "Codes Rejet", "Statut",
        "Émissions Oracle", "Tranches EDF"], COULEUR_ENTETE)

    for i, r in enumerate(rapprochement):
        lg = 4 + i
        ws2.cell(row=lg, column=1, value=r["iban_creancier"])
        ws2.cell(row=lg, column=2, value=r["echeance"].strftime("%d/%m/%Y"))
        ws2.cell(row=lg, column=3, value=r["nb_oracle"]).number_format = FORMAT_NOMBRE
        ws2.cell(row=lg, column=4, value=_f(r["montant_oracle"])).number_format = FORMAT_MONTANT
        ws2.cell(row=lg, column=5, value=r["nb_edf"]).number_format = FORMAT_NOMBRE
        ws2.cell(row=lg, column=6, value=_f(r["montant_edf"])).number_format = FORMAT_MONTANT
        ws2.cell(row=lg, column=7, value=r["ecart_nb"]).number_format = FORMAT_NOMBRE
        ws2.cell(row=lg, column=8, value=_f(r["ecart_montant"])).number_format = FORMAT_MONTANT
        ws2.cell(row=lg, column=9, value=r["nb_rejets"]).number_format = FORMAT_NOMBRE
        ws2.cell(row=lg, column=10, value=_f(r["montant_rejets"])).number_format = FORMAT_MONTANT
        ws2.cell(row=lg, column=11, value=r["codes_rejets"])
        ws2.cell(row=lg, column=12, value=r["statut"])
        ws2.cell(row=lg, column=13, value=r["emissions"])
        ws2.cell(row=lg, column=14, value=r["tranches_edf"])

        if r["statut"] in STATUTS_ANOMALIE:
            for col in (7, 8, 12):
                ws2.cell(row=lg, column=col).fill = PatternFill("solid", fgColor=COULEUR_ECART)
            ws2.cell(row=lg, column=12).font = Font(bold=True, size=TAILLE_BASE)
        elif r["statut"] in STATUTS_SIGNALES:
            ws2.cell(row=lg, column=12).fill = PatternFill("solid", fgColor=COULEUR_SIGNALE)
        elif r["statut"] == EN_ATTENTE:
            ws2.cell(row=lg, column=12).fill = PatternFill("solid", fgColor=COULEUR_ATTENTE)
        elif r["statut"] in (RAPPROCHE, EXPLIQUE_PAR_REJET, REJETE_INTEGRALEMENT):
            ws2.cell(row=lg, column=12).fill = PatternFill("solid", fgColor=COULEUR_OK)

    _colonnes_texte(ws2, (1, 2, 13, 14))
    _ajuster_largeurs(ws2, {"A": 30, "M": 24, "N": 24}, maxi=40)

    # --- Rejets ---
    ws3 = wb.create_sheet("Rejets")
    _titre(ws3, "A1", "REJETS INTERNES (APPARIÉS SUR LE RUM)")
    _entetes(ws3, 3, 1, ["IBAN Créancier", "RUM", "IBAN Débiteur", "Échéance",
                         "Montant (€)", "Code", "Motif", "Apparié Oracle",
                         "Fichier"], COULEUR_ENTETE)
    for i, r in enumerate(sorted(rejets, key=lambda x: (x.echeance, x.iban_creancier))):
        lg = 4 + i
        ws3.cell(row=lg, column=1, value=r.iban_creancier)
        ws3.cell(row=lg, column=2, value=r.rum)
        ws3.cell(row=lg, column=3, value=r.iban_debiteur)
        ws3.cell(row=lg, column=4, value=r.echeance.strftime("%d/%m/%Y"))
        ws3.cell(row=lg, column=5, value=_f(r.montant)).number_format = FORMAT_MONTANT
        ws3.cell(row=lg, column=6, value=r.code)
        ws3.cell(row=lg, column=7, value=r.motif)
        ws3.cell(row=lg, column=8, value="OUI" if r.appariee else "NON (autre SI)")
        ws3.cell(row=lg, column=9, value=r.fichier)
    _colonnes_texte(ws3, (1, 2, 3, 4))
    _ajuster_largeurs(ws3, {"A": 30, "B": 38, "C": 30}, maxi=40)

    # --- Lignes non conformes (uniquement si necessaire) ---
    if lignes_ko:
        ws4 = wb.create_sheet("Lignes rejetées")
        _titre(ws4, "A1", "LIGNES ORACLE NON CONFORMES — NON PRISES EN COMPTE")
        _entetes(ws4, 3, 1, ["Fichier", "Ligne", "Motif", "Extrait"], COULEUR_ENTETE)
        for i, k in enumerate(lignes_ko):
            lg = 4 + i
            ws4.cell(row=lg, column=1, value=k["fichier"])
            ws4.cell(row=lg, column=2, value=k["ligne"]).number_format = FORMAT_NOMBRE
            ws4.cell(row=lg, column=3, value=k["motif"])
            ws4.cell(row=lg, column=4, value=k["extrait"]).number_format = FORMAT_TEXTE
        _ajuster_largeurs(ws4, {"D": 100}, maxi=100)

    # Ecriture sous nom temporaire puis renommage : jamais de fichier a moitie
    # ecrit si le disque ou Excel pose probleme.
    temporaire = chemin.with_suffix(".tmp.xlsx")
    wb.save(temporaire)
    os.replace(temporaire, chemin)


def generer_csv(chemin, rapprochement):
    colonnes = ["iban_creancier", "echeance", "nb_oracle", "montant_oracle",
                "nb_edf", "montant_edf", "ecart_nb", "ecart_montant",
                "nb_rejets", "montant_rejets", "codes_rejets", "statut",
                "emissions", "tranches_edf"]
    temporaire = chemin.with_suffix(".tmp.csv")
    with temporaire.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=colonnes, delimiter=";")
        writer.writeheader()
        for r in rapprochement:
            ligne = dict(r)
            ligne["echeance"] = r["echeance"].strftime("%d/%m/%Y")
            writer.writerow({c: ligne[c] for c in colonnes})
    os.replace(temporaire, chemin)


# ---------------------------------------------------------------------------
# Point d'entree
# ---------------------------------------------------------------------------
def construire_parser():
    p = argparse.ArgumentParser(
        description="Rapprochement des prelevements Oracle / EDF par cle metier.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    p.add_argument("--racine", type=Path, default=Path(__file__).resolve().parent,
                   help="Racine de traitement")
    p.add_argument("--sortie", type=Path, default=None,
                   help="Dossier de sortie (defaut : la racine)")
    p.add_argument("--date", default=None,
                   help="Date de reference AAAA-MM-JJ (defaut : aujourd'hui)")
    p.add_argument("--jours", type=int, default=10,
                   help="Profondeur du perimetre, en jours calendaires")
    p.add_argument("--dossier-oracle", default="ORACLE")
    p.add_argument("--dossier-edf", default="EDF")
    p.add_argument("--motifs-oracle", nargs="+", default=["*PCX*", "*PCL*"])
    p.add_argument("--motif-edf", default="IMPORT_AVP_DK.*.*.csv")
    p.add_argument("--motif-rejets", default="REJETS_INTERNES_DK.*.csv")
    p.add_argument("--nom-si", default="ORACLE")
    return p


def trouver_dossier_rejets(edf_path):
    if edf_path.is_dir():
        for d in edf_path.iterdir():
            if d.is_dir() and d.name.upper() == "REJETS":
                return d
    return edf_path / "REJETS"


def analyser(args, diag):
    """Retourne (rapprochement, rejets, lignes_ko, contexte)."""
    racine = args.racine.resolve()
    if not racine.is_dir():
        raise ErreurTraitement(f"Racine de traitement introuvable : {racine}")

    if args.date:
        try:
            reference = datetime.strptime(args.date, "%Y-%m-%d").date()
        except ValueError:
            raise ErreurTraitement(f"Date de reference invalide : {args.date!r} (attendu AAAA-MM-JJ)")
    else:
        reference = date.today()
    if args.jours < 1:
        raise ErreurTraitement("--jours doit valoir au moins 1.")

    edf_path = racine / args.dossier_edf

    lignes_oracle, lignes_ko = charger_oracle(
        racine / args.dossier_oracle, args.motifs_oracle, diag)
    print(f"Oracle : {diag.fichiers_oracle} fichier(s), {diag.lignes_oracle} ligne(s)"
          + (f", {len(lignes_ko)} NON CONFORME(S)" if lignes_ko else ""))

    agregats_edf, dates_edf = charger_edf(edf_path, args.motif_edf, args.nom_si, diag)
    print(f"EDF : {diag.fichiers_edf} fichier(s), {diag.lignes_edf} ligne(s) '{args.nom_si}'"
          + (f", dont {diag.fichiers_edf_sans_si} fichier(s) sans ligne '{args.nom_si}'"
             if diag.fichiers_edf_sans_si else "") + ".")

    rejets = charger_rejets(trouver_dossier_rejets(edf_path), args.motif_rejets, diag)
    print(f"Rejets : {diag.fichiers_rejets} fichier(s), {diag.lignes_rejets} ligne(s)"
          + (f", {diag.rejets_dupliques} doublon(s) ecarte(s)" if diag.rejets_dupliques else "")
          + ".")

    rejets_par_cle = apparier_rejets(lignes_oracle, rejets)

    # Un flux EDF interrompu ferait basculer toutes les cles en NON_RECU.
    if dates_edf:
        ecart = (reference - max(dates_edf)).days
        if ecart >= 3:
            diag.avertir(f"Aucun fichier EDF depuis {max(dates_edf)} ({ecart} jours) : "
                         "les statuts NON_RECU sont probablement dus a une interruption du flux.")
    else:
        diag.avertir("Aucun fichier EDF exploitable : le rapprochement est non concluant.")

    couverture = min((lo.emission for lo in lignes_oracle), default=None)
    if couverture and couverture > reference - timedelta(days=args.jours):
        diag.avertir(f"L'historique Oracle demarre le {couverture}, apres le debut du "
                     f"perimetre demande ({reference - timedelta(days=args.jours)}) : "
                     "resultat partiel.")

    rapprochement = construire_rapprochement(
        lignes_oracle, agregats_edf, rejets_par_cle, dates_edf, reference, args.jours)

    contexte = [
        ("Date de référence", reference.strftime("%d/%m/%Y")),
        ("Profondeur (jours)", args.jours),
        ("Échéances retenues", f"{(reference - timedelta(days=args.jours)).strftime('%d/%m/%Y')}"
                               f" → {(reference + timedelta(days=MARGE_ECHEANCE_FUTURE)).strftime('%d/%m/%Y')}"),
        ("Lignes Oracle", diag.lignes_oracle),
        ("Lignes EDF", diag.lignes_edf),
        ("Rejets retenus", diag.lignes_rejets),
        ("Rejets appariés Oracle", sum(1 for r in rejets if r.appariee)),
        ("Clés rapprochées", len(rapprochement)),
    ]
    return rapprochement, rejets, lignes_ko, contexte, reference


def main(argv=None):
    args = construire_parser().parse_args(argv)
    diag = Diagnostic()

    try:
        rapprochement, rejets, lignes_ko, contexte, reference = analyser(args, diag)

        sortie = (args.sortie or args.racine).resolve()
        sortie.mkdir(parents=True, exist_ok=True)
        base = f"Rapprochement_Cle_Metier_{reference.strftime('%Y%m%d')}_" \
               f"{datetime.now().strftime('%H%M%S')}"

        resume = {}
        for r in rapprochement:
            e = resume.setdefault(r["statut"], {"cles": 0, "nb": 0, "montant": Decimal(0)})
            e["cles"] += 1
            e["nb"] += r["nb_oracle"] or r["nb_edf"]
            e["montant"] += r["montant_oracle"] or r["montant_edf"]

        print("Génération du rapport...")
        generer_classeur(sortie / f"{base}.xlsx", rapprochement, rejets,
                         lignes_ko, resume, contexte)
        generer_csv(sortie / f"{base}.csv", rapprochement)

    except ErreurTraitement as exc:
        print(f"Erreur critique : {exc}", file=sys.stderr)
        return 2
    except OSError as exc:
        print(f"Erreur d'acces fichier : {exc}", file=sys.stderr)
        return 2

    anomalies = sum(1 for r in rapprochement if r["statut"] in STATUTS_ANOMALIE)
    signales = sum(1 for r in rapprochement if r["statut"] in STATUTS_SIGNALES)

    print("\n=======================================================")
    print(f" Rapport : {sortie / (base + '.xlsx')}")
    for statut in ORDRE_STATUTS:
        if statut in resume:
            print(f"   {resume[statut]['cles']:5}  {statut}")
    print("-------------------------------------------------------")
    if lignes_ko:
        print(f" {len(lignes_ko)} ligne(s) Oracle non conforme(s) — résultat non fiable.")
    if signales:
        print(f" {signales} clé(s) à signaler au métier.")
    if anomalies:
        print(f" {anomalies} ANOMALIE(S) à traiter.")
    else:
        print(" Aucune anomalie : tout est rapproché ou expliqué.")
    if diag.avertissements:
        print(f" {len(diag.avertissements)} avertissement(s) — exécution dégradée.")
    print("=======================================================")

    if lignes_ko:
        return 2
    if diag.avertissements:
        return 3
    return 1 if anomalies else 0


if __name__ == "__main__":
    sys.exit(main())
