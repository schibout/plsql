#!/usr/bin/env python3
"""Rapprochement des ordres de prelevement emis par ORACLE avec les etats de
reception EDF, avec restitution Excel en 3 onglets.

Etape 1 : synthese journaliere brute (volumes et montants Oracle vs EDF).
Etape 2 : etat comparatif et ecarts, avec projection de la date Oracle vers
          la date de prise en charge EDF attendue.
Etape 3 : contenu brut des fichiers de rejets internes.

Portage du script PowerShell Prelevements_Rapprochement_Oracle_EDF.ps1.
Contrairement a celui-ci, ne depend pas d'une installation locale de Microsoft
Excel : la generation du .xlsx passe par openpyxl.

Usage : python prelevements_rapprochement.py [--racine .] [--sortie .]

Code retour : 0 = aucun ecart, 1 = ecarts detectes, 2 = erreur de traitement.
"""
import argparse
import fnmatch
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - message d'aide a l'installation
    sys.exit("Le module 'openpyxl' est requis. Installez-le avec :\n"
             "    python -m pip install -r requirements.txt")

# Encodage des fichiers sources (ISO-8859-1 : les etats EDF et Oracle sont
# produits en ANSI/CP1252, dont les accents francais coincident sur cette plage).
ENCODAGE_SOURCE = "latin-1"

# Couleurs telles que rendues par le script PowerShell d'origine.
# Attention : celui-ci passe les couleurs a Excel en BGR ; les valeurs ci-dessous
# sont les RGB reellement obtenus, pour que les deux scripts produisent le meme
# classeur a l'octet pres cote mise en forme.
COULEUR_ENTETE = "FF1F497D"   # bleu fonce
COULEUR_ECART = "FFFCE4D6"    # peche : ligne en ecart
COULEUR_OK = "FFEDEFDA"       # vert pale : ligne conforme
COULEUR_BRUT = "FFBD814F"     # brun : entete de l'onglet des rejets

FORMAT_MONTANT = "#,##0.00"
FORMAT_NOMBRE = "#,##0"
FORMAT_TEXTE = "@"

BLANC = "FFFFFFFF"
TAILLE_BASE = 11   # taille ecrite explicitement, comme le fait Excel via COM

# Tous les fichiers produits atterrissent ici, sous le dossier de sortie.
DOSSIER_RAPPORT = "rapport"


class ErreurTraitement(Exception):
    """Erreur fonctionnelle : interrompt le traitement avec un message clair."""


class Diagnostic:
    """Compteurs de traitement : un rapport partiel doit se voir."""

    def __init__(self):
        self.fichiers_oracle = 0
        self.lignes_oracle = 0
        self.lignes_oracle_ignorees = 0
        self.fichiers_edf = 0
        self.lignes_edf = 0
        self.fichiers_rejets = 0
        self.lignes_rejets = 0
        self.avertissements = []

    def avertir(self, message):
        self.avertissements.append(message)
        print(f"AVERTISSEMENT : {message}", file=sys.stderr)


def lire_lignes(chemin):
    """Lecture deterministe : jamais de surprise selon la locale du poste."""
    return chemin.read_text(encoding=ENCODAGE_SOURCE).splitlines()


def vers_decimal(valeur):
    """Conversion invariante : le separateur decimal est toujours le point."""
    return Decimal(valeur.strip().replace(",", "."))


def parse_date(valeur):
    """Renvoie None au lieu de lever, pour pouvoir signaler l'anomalie."""
    try:
        return datetime.strptime(valeur, "%Y%m%d")
    except (ValueError, TypeError):
        return None


def index_colonne_montant(lignes):
    """Resout la position de la colonne AMOUNT depuis la ligne de description
    FORMAT1ENTITYID plutot que de la coder en dur : le format Oracle peut
    evoluer sans que le script ne s'en apercoive."""
    for ligne in lignes:
        if ligne.startswith("FORMAT1ENTITYID"):
            colonnes = [c.strip() for c in ligne.split(",")]
            if "AMOUNT" in colonnes:
                return colonnes.index("AMOUNT")
    return -1


# ---------------------------------------------------------------------------
# 2. Analyse ORACLE
# ---------------------------------------------------------------------------
def analyser_oracle(oracle_path, motifs, diag):
    """Agrege nombre de lignes et montant total par date de generation
    (le nom du sous-dossier)."""
    if not oracle_path.is_dir():
        raise ErreurTraitement(f"Dossier Oracle introuvable : {oracle_path}")

    brut = defaultdict(lambda: {"nb": 0, "total": Decimal(0), "fichiers": set()})

    for fichier in sorted(oracle_path.rglob("*")):
        if not fichier.is_file():
            continue
        if not any(fnmatch.fnmatch(fichier.name, m) for m in motifs):
            continue

        date_dossier = fichier.parent.name
        if parse_date(date_dossier) is None:
            diag.avertir(f"Dossier non date ignore : {fichier.parent}")
            continue

        lignes = lire_lignes(fichier)
        idx_montant = index_colonne_montant(lignes)
        if idx_montant < 0:
            diag.avertir(f"Colonne AMOUNT introuvable, index 5 par defaut : {fichier.name}")
            idx_montant = 5

        diag.fichiers_oracle += 1
        for ligne in lignes:
            if not ligne.strip():
                continue
            if ligne.startswith("HEADER") or ligne.startswith("FORMAT1ENTITYID"):
                continue

            champs = ligne.split(",")
            if len(champs) <= idx_montant:
                diag.lignes_oracle_ignorees += 1
                continue
            try:
                montant = vers_decimal(champs[idx_montant])
            except InvalidOperation:
                diag.lignes_oracle_ignorees += 1
                continue

            brut[date_dossier]["nb"] += 1
            brut[date_dossier]["total"] += montant
            brut[date_dossier]["fichiers"].add(fichier.name)
            diag.lignes_oracle += 1

    if diag.lignes_oracle_ignorees:
        diag.avertir(f"{diag.lignes_oracle_ignorees} ligne(s) Oracle illisible(s) "
                     "ecartee(s) du calcul.")
    if not brut:
        raise ErreurTraitement(f"Aucune ligne Oracle exploitable trouvee sous {oracle_path}.")

    return [{"date": d, "nb": v["nb"], "total": v["total"],
             "fichiers": sorted(v["fichiers"])} for d, v in sorted(brut.items())]


# ---------------------------------------------------------------------------
# 3. Analyse EDF
# ---------------------------------------------------------------------------
def analyser_edf(edf_path, motif, nom_si, diag):
    """Agrege les lignes du SI cible par date de reception (nom du fichier)."""
    if not edf_path.is_dir():
        diag.avertir(f"Dossier EDF introuvable : {edf_path}. "
                     "Les volumes recus seront a zero.")
        return []

    brut = defaultdict(lambda: {"nb": 0, "total": Decimal(0), "fichiers": set()})

    for fichier in sorted(edf_path.glob(motif)):
        if not fichier.is_file():
            continue
        segments = fichier.name.split(".")
        if len(segments) < 2 or parse_date(segments[1]) is None:
            diag.avertir(f"Date illisible dans le nom du fichier EDF, ignore : {fichier.name}")
            continue

        date_edf = segments[1]
        diag.fichiers_edf += 1

        for ligne in lire_lignes(fichier):
            champs = ligne.split(";")
            if len(champs) < 5 or champs[0].strip() != nom_si:
                continue
            try:
                nb = int(champs[3].strip())
                total = vers_decimal(champs[4])
            except (ValueError, InvalidOperation):
                diag.avertir(f"Ligne EDF illisible ignoree dans {fichier.name} : {ligne}")
                continue

            brut[date_edf]["nb"] += nb
            brut[date_edf]["total"] += total
            brut[date_edf]["fichiers"].add(fichier.name)
            diag.lignes_edf += 1

    # Un rapport ou tout l'EDF est a zero ressemble a un ecart total : il faut
    # distinguer "rien recu" de "mauvais filtre / format change".
    if diag.fichiers_edf and not diag.lignes_edf:
        diag.avertir(f"Aucune ligne '{nom_si}' trouvee dans les {diag.fichiers_edf} "
                     "fichier(s) EDF lus. Verifiez --nom-si ou le format des fichiers : "
                     "les ecarts affiches seront trompeurs.")

    return [{"date": d, "nb": v["nb"], "total": v["total"],
             "fichiers": sorted(v["fichiers"])} for d, v in sorted(brut.items())]


# ---------------------------------------------------------------------------
# 4. Lecture brute des fichiers de REJETS
# ---------------------------------------------------------------------------
PREFIXE_TITRE_REJETS = "LISTE DES REJETS INTERNES AU"


def lire_rejets_structures(rejets_path, motif, diag):
    """Decoupe les fichiers de rejets en tableau exploitable.

    Structure constante des fichiers (verifiee sur l'ensemble du gisement) :
      ligne 1 : 'LISTE DES REJETS INTERNES AU:<date> a <heure>'  -> devient une COLONNE
      ligne 2 : l'en-tete des champs                             -> devient les COLONNES
      ligne 3 et suivantes : les donnees                         -> deviennent les LIGNES

    L'en-tete est lu DANS le fichier, jamais code en dur : un changement de
    format doit se voir plutot que de decaler silencieusement les colonnes.

    Retourne (colonnes, lignes).
    """
    lignes = []
    colonnes = []
    if not rejets_path.is_dir():
        diag.avertir(f"Dossier de rejets introuvable : {rejets_path}.")
        return colonnes, lignes

    fichiers = sorted(p for p in rejets_path.iterdir()
                      if p.is_file() and fnmatch.fnmatch(p.name, motif))
    diag.fichiers_rejets = len(fichiers)

    for fichier in fichiers:
        contenu = [l for l in lire_lignes(fichier) if l.strip()]
        if len(contenu) < 2:
            diag.avertir(f"Fichier de rejets trop court, ignore : {fichier.name}")
            continue

        titre, entete = contenu[0], contenu[1]
        if not titre.startswith(PREFIXE_TITRE_REJETS):
            diag.avertir(f"Titre inattendu dans {fichier.name} : {titre[:60]}")
        # Le titre porte la date d'edition apres le ':' ; c'est la seule partie
        # variable, donc la seule qui merite une colonne.
        edite_le = titre.split(":", 1)[1].strip() if ":" in titre else titre

        # Le ';' final produit un champ vide en fin d'en-tete : on l'ecarte.
        champs_entete = [c.strip() for c in entete.split(";")]
        while champs_entete and not champs_entete[-1]:
            champs_entete.pop()

        if not colonnes:
            colonnes = champs_entete
        elif champs_entete != colonnes:
            diag.avertir(f"En-tete different dans {fichier.name} : "
                         f"{entete[:80]} — colonnes possiblement decalees.")

        for brute in contenu[2:]:
            valeurs = [v.strip() for v in brute.split(";")][:len(colonnes)]
            valeurs += [""] * (len(colonnes) - len(valeurs))
            lignes.append({"fichier": fichier.name, "edite_le": edite_le,
                           "valeurs": valeurs})
            diag.lignes_rejets += 1

    return colonnes, lignes


# ---------------------------------------------------------------------------
# 5. Preparation du rapprochement comparatif
# ---------------------------------------------------------------------------
def date_cible(date_oracle):
    """Projection Oracle -> EDF : J+2 ouvre approxime en jours calendaires.

    A revoir avec le metier : ne tient pas compte des jours feries.
    weekday() : 3 = jeudi, 4 = vendredi, 5 = samedi.
    """
    if date_oracle.weekday() in (3, 4):
        return date_oracle + timedelta(days=4)
    return date_oracle + timedelta(days=2)


def resumer_rejets(rejets_path, motif, diag):
    """Resume chaque fichier de rejets : total et codes, sans toucher a l'onglet brut.

    Sert uniquement a designer le fichier a ouvrir quand un ecart apparait ;
    l'onglet 3 continue de recopier les lignes telles quelles.
    """
    resume = {}
    if not rejets_path.is_dir():
        return resume

    for fichier in sorted(p for p in rejets_path.iterdir()
                          if p.is_file() and fnmatch.fnmatch(p.name, motif)):
        segments = fichier.name.split(".")
        date_rejet = parse_date(segments[1]) if len(segments) > 1 else None
        if date_rejet is None:
            diag.avertir(f"Date illisible dans le nom du fichier de rejets : {fichier.name}")
            continue

        total, nb, codes = Decimal(0), 0, set()
        for ligne in lire_lignes(fichier):
            champs = ligne.split(";")
            # Une ligne de donnees commence par un IBAN creancier.
            if len(champs) < 7 or not champs[0].strip().startswith("FR"):
                continue
            try:
                total += vers_decimal(champs[4])
            except InvalidOperation:
                continue
            nb += 1
            codes.add(champs[5].strip())

        if nb:
            resume[date_rejet] = {"fichier": fichier.name, "total": total,
                                  "nb": nb, "codes": sorted(codes)}
    return resume


def rejets_de_la_periode(date_ora, date_edf, resume_rejets, ecart_montant):
    """Designe le ou les fichiers de rejets susceptibles d'expliquer un ecart.

    Les fichiers de rejets ne portent pas le nom du SI : la seule fenetre de
    dates ramene aussi des rejets etrangers a la ligne (constate sur 2 cas sur 6).
    On ne conclut donc que lorsque les montants concordent, et on le dit
    clairement sinon plutot que de laisser croire a une explication.
    """
    candidats = [v for d, v in sorted(resume_rejets.items()) if date_ora < d <= date_edf]
    if not candidats:
        return "(aucun)"

    manque = -ecart_montant   # un ecart negatif signifie un manque cote EDF
    if manque > 0:
        for c in candidats:
            if c["total"] == manque:
                return (f"{c['fichier']} — {c['nb']} rejet(s), "
                        f"{c['total']} € ({', '.join(c['codes'])}) : explique l'écart")
        if sum(c["total"] for c in candidats) == manque:
            return (" + ".join(c["fichier"] for c in candidats)
                    + f" — {sum(c['nb'] for c in candidats)} rejet(s) : expliquent l'écart")

    return (" + ".join(c["fichier"] for c in candidats)
            + f" — {sum(c['nb'] for c in candidats)} rejet(s), "
              f"{sum(c['total'] for c in candidats)} € : à vérifier")


def chercher_date_cible(date_ora, nb_ora, total_ora, edf_par_date, n_max):
    """Cherche la date EDF qui repond a une date Oracle, au lieu de la deviner.

    Le decalage reel n'est pas constant : il depend du jour de la semaine, des
    jours feries et du rythme de production d'EDF. Plutot que de le postuler, on
    essaie J+1 a J+n et on retient le meilleur candidat.

    Regle de selection, deduite du fonctionnement du flux : EDF remonte NET des
    rejets, donc un candidat credible ne peut jamais afficher PLUS que ce
    qu'Oracle a emis, ni en nombre ni en montant. Cette contrainte de signe est
    ce qui elimine les coincidences : une journee EDF affichant le meme NOMBRE
    de prelevements mais un montant sans rapport est ecartee.

    Retourne (date_cible AAAAMMJJ ou None, mode) ou mode vaut
    'exact', 'ecart', 'ambigu' ou 'introuvable'.
    """
    exacts, plausibles = [], []
    for n in range(1, n_max + 1):
        cible = date_ora + timedelta(days=n)
        candidat = edf_par_date.get(cible.strftime("%Y%m%d"))
        if not candidat or candidat["nb"] == 0:
            continue
        manque_nb = nb_ora - candidat["nb"]
        manque_montant = total_ora - candidat["total"]
        if (manque_nb, manque_montant) == (0, Decimal(0)):
            exacts.append((n, cible))
        elif manque_nb >= 0 and manque_montant >= 0:
            plausibles.append((manque_nb, n, cible))

    if len(exacts) == 1:
        return exacts[0][1].strftime("%Y%m%d"), "exact"
    if len(exacts) > 1:
        # Deux journees EDF concordent parfaitement : on prend la plus proche,
        # mais le rapport doit dire que le choix n'est pas certain.
        return exacts[0][1].strftime("%Y%m%d"), "ambigu"
    if plausibles:
        # Departage par l'ECART LE PLUS FAIBLE, pas par la date la plus proche :
        # une petite journee EDF sans rapport satisfait la contrainte de signe et
        # serait retenue a tort si l'on prenait simplement le premier candidat.
        plausibles.sort()
        return plausibles[0][2].strftime("%Y%m%d"), "ecart"
    return None, "introuvable"


def construire_rapprochement(oracle_summary, edf_summary, recherche_jn=0,
                             resume_rejets=None):
    edf_par_date = {e["date"]: e for e in edf_summary}

    # Plusieurs dates Oracle peuvent viser la meme date EDF : on les regroupe.
    groupes = defaultdict(list)
    modes = {}
    for ora in oracle_summary:
        date_ora = parse_date(ora["date"])
        if recherche_jn:
            cible, mode = chercher_date_cible(
                date_ora, ora["nb"], ora["total"], edf_par_date, recherche_jn)
            if cible is None:
                # Aucun candidat : on garde la projection par defaut pour que la
                # journee figure quand meme au rapport, en ecart total.
                cible = date_cible(date_ora).strftime("%Y%m%d")
            modes[cible] = mode if cible not in modes else "groupe"
        else:
            cible = date_cible(date_ora).strftime("%Y%m%d")
        groupes[cible].append(ora)

    lignes = []
    for date_edf_cible in sorted(groupes):
        groupe = groupes[date_edf_cible]
        dates_origine = sorted(o["date"] for o in groupe)

        nb_attendu = sum(o["nb"] for o in groupe)
        total_attendu = sum((o["total"] for o in groupe), Decimal(0))

        correspondance = edf_par_date.get(date_edf_cible)
        nb_recu = correspondance["nb"] if correspondance else 0
        total_recu = correspondance["total"] if correspondance else Decimal(0)

        # Provenance : de quel fichier vient chaque cote de la comparaison.
        # EDF ne produit qu'un fichier par date, son nom est donc toujours utile.
        # Oracle peut en compter jusqu'a 78 pour une seule date : au-dela de
        # trois, la liste devient illisible et le dossier suffit a s'y retrouver.
        fichiers_ora = sorted({f for o in groupe for f in o.get("fichiers", [])})
        if not fichiers_ora:
            provenance_ora = "(aucun)"
        elif len(fichiers_ora) <= 3:
            provenance_ora = " + ".join(fichiers_ora)
        else:
            dossiers = " + ".join(dates_origine)
            provenance_ora = f"ORACLE\\{dossiers} ({len(fichiers_ora)} fichiers)"
        provenance_edf = (" + ".join(correspondance.get("fichiers", []))
                          if correspondance else "(aucun)")

        # Rejets : renseignes uniquement sur les lignes en ecart, sinon la
        # colonne se remplirait de bruit sur les journees conformes.
        ecart_montant_ligne = total_recu - total_attendu
        # `is not None` et non `if resume_rejets` : un dossier de rejets vide doit
        # afficher « (aucun) » sur une ligne en ecart, l'information est utile.
        if resume_rejets is not None and ecart_montant_ligne != 0:
            provenance_rejet = rejets_de_la_periode(
                parse_date(dates_origine[0]), parse_date(date_edf_cible),
                resume_rejets, ecart_montant_ligne)
        else:
            provenance_rejet = ""

        date_edf_parsee = parse_date(date_edf_cible)
        # Delai mesure depuis la premiere date d'origine du groupe.
        delai = (date_edf_parsee - parse_date(dates_origine[0])).days
        ecart_nb = nb_recu - nb_attendu

        statut = f"OK (J+{delai})" if ecart_nb == 0 else f"Écart (J+{delai})"
        mode = modes.get(date_edf_cible)
        if mode == "ambigu":
            statut += " ambigu"
        elif mode == "introuvable":
            statut = "Non trouvé"

        lignes.append({
            "date_oracle": " + ".join(parse_date(d).strftime("%d/%m/%Y") for d in dates_origine),
            "date_edf": date_edf_parsee.strftime("%d/%m/%Y"),
            "nb_oracle": nb_attendu,
            "nb_edf": nb_recu,
            "ecart_nb": ecart_nb,
            "total_oracle": arrondi(total_attendu),
            "total_edf": arrondi(total_recu),
            "ecart_montant": arrondi(total_recu - total_attendu),
            "statut": statut,
            "fichier_oracle": provenance_ora,
            "fichier_edf": provenance_edf,
            "fichier_rejet": provenance_rejet,
        })

    return lignes


def arrondi(valeur):
    return float(round(valeur, 2))


# ---------------------------------------------------------------------------
# 6. Restitution Excel
# ---------------------------------------------------------------------------
def _titre(ws, cellule, texte):
    ws[cellule] = texte
    ws[cellule].font = Font(bold=True, size=14)


def _entetes(ws, ligne, col_depart, libelles, couleur):
    fill = PatternFill("solid", fgColor=couleur)
    for i, libelle in enumerate(libelles):
        cell = ws.cell(row=ligne, column=col_depart + i, value=libelle)
        cell.fill = fill
        cell.font = Font(bold=True, color=BLANC, size=TAILLE_BASE)
        cell.alignment = Alignment(horizontal="center")


def _colonnes_texte(ws, colonnes):
    """Force le format texte sur les cellules renseignees des colonnes visees.

    Reproduit le comportement de PowerShell, qui applique le format a la colonne
    entiere : dates jj/mm/aaaa, IBAN et zeros de tete ne sont jamais reinterpretes
    par Excel. Les cellules vides sont laissees telles quelles, comme dans le
    classeur produit par le script PowerShell.
    """
    for col in colonnes:
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.number_format = FORMAT_TEXTE


def _ajuster_largeurs(ws, largeurs_imposees=None, maxi=60):
    """openpyxl n'a pas d'equivalent a AutoFit : on approxime a partir du contenu."""
    largeurs = defaultdict(int)
    for ligne in ws.iter_rows():
        for cell in ligne:
            if cell.value is not None:
                largeurs[cell.column] = max(largeurs[cell.column], len(str(cell.value)))
    for col, largeur in largeurs.items():
        ws.column_dimensions[get_column_letter(col)].width = min(largeur + 2, maxi)
    for col, largeur in (largeurs_imposees or {}).items():
        ws.column_dimensions[col].width = largeur


def generer_classeur(chemin, oracle_summary, edf_summary, rapprochement, rejets):
    wb = Workbook()

    # --- ONGLET 1 : Synthese journaliere brute ---
    ws1 = wb.active
    ws1.title = "Synthèse Journalière Brute"
    _titre(ws1, "A1", "ÉTAPE 1 : SYNTHÈSE JOURNALIÈRE BRUTE")
    ws1["A3"] = "Données Émises par ORACLE"
    ws1["A3"].font = Font(bold=True, size=TAILLE_BASE)
    ws1["E3"] = "Données Reçues par EDF"
    ws1["E3"].font = Font(bold=True, size=TAILLE_BASE)

    _entetes(ws1, 4, 1, ["Date Gén. Oracle", "Nb Total Lignes", "Montant Total (€)"],
             COULEUR_ENTETE)
    _entetes(ws1, 4, 5, ["Date Prise Chg EDF", "Nb Prél. Pris en Chg",
                         "Montant Pris en Chg (€)"], COULEUR_ENTETE)

    for i, row in enumerate(oracle_summary):
        r = 5 + i
        # Les dates sont stockees en texte pour figer l'affichage jj/mm/aaaa.
        c = ws1.cell(row=r, column=1, value=parse_date(row["date"]).strftime("%d/%m/%Y"))
        c.number_format = FORMAT_TEXTE
        c.alignment = Alignment(horizontal="center")
        ws1.cell(row=r, column=2, value=row["nb"]).number_format = FORMAT_NOMBRE
        ws1.cell(row=r, column=3, value=arrondi(row["total"])).number_format = FORMAT_MONTANT

    for i, row in enumerate(edf_summary):
        r = 5 + i
        c = ws1.cell(row=r, column=5, value=parse_date(row["date"]).strftime("%d/%m/%Y"))
        c.number_format = FORMAT_TEXTE
        c.alignment = Alignment(horizontal="center")
        ws1.cell(row=r, column=6, value=row["nb"]).number_format = FORMAT_NOMBRE
        ws1.cell(row=r, column=7, value=arrondi(row["total"])).number_format = FORMAT_MONTANT

    _colonnes_texte(ws1, (1, 5))
    _ajuster_largeurs(ws1, {"A": 26})

    # --- ONGLET 2 : Etat comparatif et ecarts ---
    ws2 = wb.create_sheet("État Comparatif & Écarts")
    _titre(ws2, "A1", "ÉTAPE 2 : ÉTAT COMPARATIF ET RAPPROCHEMENT DES ÉCARTS")
    _entetes(ws2, 3, 1, [
        "Date(s) Gén. Oracle", "Date Reç. EDF (Cible)", "Nb Attendu (Ora)",
        "Nb Reçu (EDF)", "Écart Nombre", "Total Attendu (Ora)", "Total Reçu (EDF)",
        "Écart Montant (€)", "Statut / Délai",
        "Fichier(s) ORACLE", "Fichier EDF", "Fichier(s) REJET"], COULEUR_ENTETE)

    fill_ecart = PatternFill("solid", fgColor=COULEUR_ECART)
    fill_ok = PatternFill("solid", fgColor=COULEUR_OK)

    for i, row in enumerate(rapprochement):
        r = 4 + i
        for col, valeur in ((1, row["date_oracle"]), (2, row["date_edf"])):
            c = ws2.cell(row=r, column=col, value=valeur)
            c.number_format = FORMAT_TEXTE
            c.alignment = Alignment(horizontal="center")
        for col, valeur in ((3, row["nb_oracle"]), (4, row["nb_edf"]), (5, row["ecart_nb"])):
            ws2.cell(row=r, column=col, value=valeur).number_format = FORMAT_NOMBRE
        for col, valeur in ((6, row["total_oracle"]), (7, row["total_edf"]),
                            (8, row["ecart_montant"])):
            ws2.cell(row=r, column=col, value=valeur).number_format = FORMAT_MONTANT
        ws2.cell(row=r, column=9, value=row["statut"])
        ws2.cell(row=r, column=10, value=row["fichier_oracle"])
        ws2.cell(row=r, column=11, value=row["fichier_edf"])
        ws2.cell(row=r, column=12, value=row["fichier_rejet"])

        if row["ecart_nb"] != 0:
            ws2.cell(row=r, column=5).fill = fill_ecart
            ws2.cell(row=r, column=8).fill = fill_ecart
            ws2.cell(row=r, column=9).fill = fill_ecart
            ws2.cell(row=r, column=9).font = Font(bold=True, size=TAILLE_BASE)
        else:
            ws2.cell(row=r, column=9).fill = fill_ok

    # Les noms de fichiers restent en texte : jamais reinterpretes par Excel.
    _colonnes_texte(ws2, (1, 2, 10, 11, 12))
    _ajuster_largeurs(ws2, {"A": 22, "J": 62, "K": 34, "L": 70}, maxi=70)

    # --- ONGLET 3 : Rejets internes, sous forme de tableau ---
    colonnes_rejets, lignes_rejets = rejets
    ws3 = wb.create_sheet("Fichiers de Rejets Bruts")
    _titre(ws3, "A1", "REJETS INTERNES")

    if not colonnes_rejets:
        ws3["A3"] = "Aucun fichier de rejets exploitable."
        ws3["A3"].font = Font(bold=True, size=TAILLE_BASE)
        ws3.column_dimensions["A"].width = 50
        wb.save(chemin)
        return

    # La 1re ligne du fichier devient une colonne, la 2e fournit les en-tetes.
    entetes = ["Nom du Fichier Source", PREFIXE_TITRE_REJETS] + colonnes_rejets
    _entetes(ws3, 3, 1, entetes, COULEUR_BRUT)

    # Le montant est ecrit en nombre pour rester sommable dans Excel ; tout le
    # reste (IBAN, RUM, dates) doit rester du texte strict.
    try:
        idx_montant_rejet = colonnes_rejets.index("MONTANT")
    except ValueError:
        idx_montant_rejet = -1

    for i, rejet in enumerate(lignes_rejets):
        r = 4 + i
        ws3.cell(row=r, column=1, value=rejet["fichier"])
        ws3.cell(row=r, column=2, value=rejet["edite_le"])
        for j, valeur in enumerate(rejet["valeurs"]):
            cell = ws3.cell(row=r, column=3 + j)
            if j == idx_montant_rejet:
                try:
                    cell.value = float(vers_decimal(valeur))
                    cell.number_format = FORMAT_MONTANT
                    continue
                except InvalidOperation:
                    pass          # montant illisible : on garde le texte brut
            cell.value = valeur

    colonnes_texte = [c for c in range(1, 3 + len(colonnes_rejets))
                      if c != 3 + idx_montant_rejet]
    _colonnes_texte(ws3, colonnes_texte)

    ws3.auto_filter.ref = f"A3:{get_column_letter(len(entetes))}{3 + len(lignes_rejets)}"
    ws3.freeze_panes = "A4"
    _ajuster_largeurs(ws3, {"A": 38, "B": 20}, maxi=38)

    wb.save(chemin)


# ---------------------------------------------------------------------------
# Point d'entree
# ---------------------------------------------------------------------------
def construire_parser():
    p = argparse.ArgumentParser(
        description="Rapprochement des prelevements Oracle / EDF.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    p.add_argument("--racine", type=Path, default=Path(__file__).resolve().parent,
                   help="Racine de traitement")
    p.add_argument("--sortie", type=Path, default=None,
                   help="Dossier de sortie (defaut : la racine). Le classeur est "
                        f"ecrit dans son sous-dossier '{DOSSIER_RAPPORT}'")
    p.add_argument("--dossier-oracle", default="ORACLE",
                   help="Sous-dossier de l'arborescence Oracle")
    p.add_argument("--dossier-edf", default="EDF",
                   help="Sous-dossier des etats de reception EDF")
    p.add_argument("--motifs-oracle", nargs="+", default=["*PCX*", "*PCL*"],
                   help="Motifs de nom des fichiers Oracle")
    p.add_argument("--motif-edf", default="IMPORT_AVP_DK.*.*.csv",
                   help="Motif de nom des etats de reception EDF")
    p.add_argument("--motif-rejets", default="REJETS_INTERNES_DK.*",
                   help="Motif de nom des fichiers de rejets internes")
    p.add_argument("--nom-si", default="ORACLE",
                   help="Valeur de la colonne 'NOM DU SI' a rapprocher")
    p.add_argument("--recherche-jn", type=int, default=0, metavar="N",
                   help="Chercher la date EDF parmi J+1..J+N au lieu d'appliquer "
                        "la regle fixe J+2/J+4. 0 = desactive. Valeur conseillee : 8")
    return p


def trouver_dossier_rejets(edf_path):
    """Detection du dossier rejets sans distinction de casse."""
    if edf_path.is_dir():
        for d in edf_path.iterdir():
            if d.is_dir() and d.name.upper() == "REJETS":
                return d
    return edf_path / "REJETS"


def main(argv=None):
    args = construire_parser().parse_args(argv)
    diag = Diagnostic()

    try:
        racine = args.racine.resolve()
        if not racine.is_dir():
            raise ErreurTraitement(f"Racine de traitement introuvable : {racine}")
        # Les rapports sont toujours regroupes dans un sous-dossier dedie : ils
        # ne se melangent jamais aux fichiers sources analyses.
        sortie = (args.sortie or racine).resolve() / DOSSIER_RAPPORT
        sortie.mkdir(parents=True, exist_ok=True)

        edf_path = racine / args.dossier_edf
        horodatage = datetime.now().strftime("%Y%m%d_%H%M%S")
        chemin_sortie = sortie / f"Rapprochement_Oracle_EDF_{horodatage}.xlsx"

        print(f"Analyse du dossier en cours : {racine.name}")

        oracle_summary = analyser_oracle(racine / args.dossier_oracle,
                                         args.motifs_oracle, diag)
        print(f"Oracle : {diag.fichiers_oracle} fichier(s), {diag.lignes_oracle} ligne(s), "
              f"{len(oracle_summary)} date(s).")

        edf_summary = analyser_edf(edf_path, args.motif_edf, args.nom_si, diag)
        if diag.fichiers_edf:
            print(f"EDF : {diag.fichiers_edf} fichier(s), {diag.lignes_edf} ligne(s) "
                  f"'{args.nom_si}', {len(edf_summary)} date(s).")

        rejets = lire_rejets_structures(trouver_dossier_rejets(edf_path),
                                        args.motif_rejets, diag)
        if diag.fichiers_rejets:
            print(f"Rejets : {diag.fichiers_rejets} fichier(s), {diag.lignes_rejets} ligne(s).")

        resume_rejets = resumer_rejets(trouver_dossier_rejets(edf_path),
                                       args.motif_rejets, diag)
        rapprochement = construire_rapprochement(oracle_summary, edf_summary,
                                                 args.recherche_jn, resume_rejets)

        print("Génération du classeur Excel...")
        generer_classeur(chemin_sortie, oracle_summary, edf_summary, rapprochement, rejets)

    except ErreurTraitement as exc:
        print(f"Erreur critique : {exc}", file=sys.stderr)
        return 2
    except OSError as exc:
        print(f"Erreur d'acces fichier : {exc}", file=sys.stderr)
        return 2

    nb_ecarts = sum(1 for r in rapprochement if r["ecart_nb"] != 0)

    print("\n=======================================================")
    print(" Rapprochement complété avec succès !")
    print(f" Fichier : {chemin_sortie}")
    print(f" Oracle : {diag.lignes_oracle} ligne(s) / EDF : {diag.lignes_edf} ligne(s)"
          f" / Rejets : {diag.lignes_rejets} ligne(s)")
    if diag.avertissements:
        print(f" {len(diag.avertissements)} avertissement(s) - rapport potentiellement incomplet.")
    if nb_ecarts:
        print(f" {nb_ecarts} date(s) en écart sur {len(rapprochement)} rapprochée(s).")
    else:
        print(f" Aucun écart sur {len(rapprochement)} date(s) rapprochée(s).")
    print("=======================================================")

    return 1 if nb_ecarts else 0


if __name__ == "__main__":
    sys.exit(main())
