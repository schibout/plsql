import json

import pytest
from openpyxl import Workbook, load_workbook

import rapport_excel


NOM_SRC = "CEL01_SRC_FACTURESFOURNISSEURS_100826-201628_ST_CEL01_1_001.csv"

COLONNES = ["Folio", "Numero Piece", "Montant Fichier", "Montant Oracle", "Statut"]
LIGNES = [
    {"Folio": "CEC", "Numero Piece": "A1", "Montant Fichier": 12.5,
     "Montant Oracle": 12.5, "Statut": "INTEGREE"},
    {"Folio": "CEE", "Numero Piece": "A2", "Montant Fichier": -3.0,
     "Montant Oracle": None, "Statut": "ABSENTE"},
]


def classeur_synthese(tmp_path, nom="FAC02_SYNTHESE_FOURNISSEURS_100826-201628.xlsx"):
    wb = Workbook()
    wb.active.title = "Synthese"
    chemin = tmp_path / nom
    wb.save(chemin)
    return chemin


def onglet(nom="Detail Oracle", colonnes=None, lignes=None):
    return {
        "nom": nom,
        "titre": "CONTROLE - DETAIL ORACLE",
        "sous_titre": "Execution : 25/08/2026 11:00:00",
        "colonnes": COLONNES if colonnes is None else colonnes,
        "colonne_statut": "Statut",
        "valeurs_ok": ["INTEGREE"],
        "lignes": LIGNES if lignes is None else lignes,
    }


def src_d_export(tmp_path, flux="10082026_FOURNISSEUR_CEG", export="c0603921"):
    source = tmp_path / flux / export / "SOURCE"
    source.mkdir(parents=True)
    src = source / NOM_SRC
    src.write_text("", encoding="latin-1")
    return src


def test_le_classeur_porte_le_nom_du_dossier_de_flux(monkeypatch, tmp_path):
    monkeypatch.setenv("CONTROLE_FLUX_RAPPORT_DIR", str(tmp_path / "rapport"))
    src = src_d_export(tmp_path)

    chemin = rapport_excel.chemin_rapport(src, "FAC02_SYNTHESE_FOURNISSEURS")

    assert chemin.endswith(
        "FAC02_SYNTHESE_FOURNISSEURS_10082026_FOURNISSEUR_CEG.xlsx"
    )


def test_le_nom_ne_depend_pas_de_l_argument_saisi(monkeypatch, tmp_path):
    """Dossier de flux, dossier d'export ou fichier : meme classeur au final."""
    monkeypatch.setenv("CONTROLE_FLUX_RAPPORT_DIR", str(tmp_path / "rapport"))
    src = src_d_export(tmp_path)

    assert rapport_excel.chemin_rapport(src, "P") == rapport_excel.chemin_rapport(
        str(src).replace("\\", "/"), "P"
    )


def test_deux_exports_du_meme_flux_donnent_le_meme_classeur(monkeypatch, tmp_path):
    monkeypatch.setenv("CONTROLE_FLUX_RAPPORT_DIR", str(tmp_path / "rapport"))
    premier = src_d_export(tmp_path, export="export1")
    second = src_d_export(tmp_path, export="export2")

    assert rapport_excel.chemin_rapport(premier, "P") == rapport_excel.chemin_rapport(
        second, "P"
    )


def test_src_hors_dossier_d_export_retombe_sur_l_horodatage(monkeypatch, tmp_path):
    monkeypatch.setenv("CONTROLE_FLUX_RAPPORT_DIR", str(tmp_path))

    chemin = rapport_excel.chemin_rapport(NOM_SRC, "FAC02_SYNTHESE_FOURNISSEURS")

    assert chemin == str(tmp_path / "FAC02_SYNTHESE_FOURNISSEURS_100826-201628.xlsx")


def test_src_sans_horodatage_ni_dossier_reste_deterministe(monkeypatch, tmp_path):
    monkeypatch.setenv("CONTROLE_FLUX_RAPPORT_DIR", str(tmp_path))

    chemin = rapport_excel.chemin_rapport("SANS_DATE.csv", "GL_SYNTHESE")

    assert chemin.endswith("GL_SYNTHESE_rapport.xlsx")


def test_les_onglets_sont_ajoutes_sans_toucher_aux_existants(tmp_path):
    chemin = classeur_synthese(tmp_path)

    rapport_excel.ajouter_onglets(chemin, [onglet()])

    wb = load_workbook(chemin)
    assert wb.sheetnames == ["Synthese", "Detail Oracle"]
    ws = wb["Detail Oracle"]
    assert ws.cell(4, 1).value == "Folio"
    assert ws.cell(5, 1).value == "CEC"
    assert ws.cell(6, 3).value == -3.0


def test_le_statut_est_colore_selon_les_valeurs_ok(tmp_path):
    chemin = classeur_synthese(tmp_path)

    rapport_excel.ajouter_onglets(chemin, [onglet()])

    ws = load_workbook(chemin)["Detail Oracle"]
    colonne_statut = COLONNES.index("Statut") + 1
    assert ws.cell(5, colonne_statut).fill.fgColor.rgb.endswith(rapport_excel.VERT)
    assert ws.cell(6, colonne_statut).fill.fgColor.rgb.endswith(rapport_excel.PECHE)


def test_les_colonnes_de_montant_sont_formatees(tmp_path):
    chemin = classeur_synthese(tmp_path)

    rapport_excel.ajouter_onglets(chemin, [onglet()])

    ws = load_workbook(chemin)["Detail Oracle"]
    assert ws.cell(5, 3).number_format == rapport_excel.FORMAT_MONTANT
    assert ws.cell(5, 1).number_format != rapport_excel.FORMAT_MONTANT


def test_une_valeur_absente_ne_bloque_pas_l_ecriture(tmp_path):
    """PowerShell peut serialiser une valeur absente en objet vide."""
    chemin = classeur_synthese(tmp_path)
    lignes = [{"Folio": "CEC", "Numero Piece": "A1", "Montant Fichier": 1.0,
               "Montant Oracle": {}, "Statut": "ABSENTE"}]

    rapport_excel.ajouter_onglets(chemin, [onglet(lignes=lignes)])

    ws = load_workbook(chemin)["Detail Oracle"]
    assert ws.cell(5, 4).value is None


def test_relancer_le_controle_remplace_l_onglet(tmp_path):
    chemin = classeur_synthese(tmp_path)
    rapport_excel.ajouter_onglets(chemin, [onglet()])

    rapport_excel.ajouter_onglets(chemin, [onglet(lignes=LIGNES[:1])])

    wb = load_workbook(chemin)
    assert wb.sheetnames.count("Detail Oracle") == 1
    assert wb["Detail Oracle"].max_row == 5


def test_un_onglet_sans_ligne_reste_lisible(tmp_path):
    chemin = classeur_synthese(tmp_path)

    rapport_excel.ajouter_onglets(chemin, [onglet(lignes=[])])

    ws = load_workbook(chemin)["Detail Oracle"]
    assert ws.cell(5, 1).value == "Aucune donnee"


@pytest.mark.parametrize("nom", ["Synthese Oracle", "Detail Oracle"])
def test_appel_en_ligne_de_commande(tmp_path, monkeypatch, capsys, nom):
    monkeypatch.setenv("CONTROLE_FLUX_RAPPORT_DIR", str(tmp_path))
    classeur = classeur_synthese(tmp_path)
    descriptif = tmp_path / "onglets.json"
    descriptif.write_text(json.dumps({
        "src": NOM_SRC,
        "prefixe": "FAC02_SYNTHESE_FOURNISSEURS",
        "onglets": [onglet(nom=nom)],
    }), encoding="utf-8")

    assert rapport_excel.main([str(descriptif)]) == 0
    assert capsys.readouterr().out.strip() == str(classeur)
    assert nom in load_workbook(classeur).sheetnames


def test_classeur_absent_est_une_erreur_explicite(tmp_path, monkeypatch, capsys):
    monkeypatch.setenv("CONTROLE_FLUX_RAPPORT_DIR", str(tmp_path))
    descriptif = tmp_path / "onglets.json"
    descriptif.write_text(json.dumps({
        "src": NOM_SRC, "prefixe": "ABSENT", "onglets": [onglet()],
    }), encoding="utf-8")

    assert rapport_excel.main([str(descriptif)]) == 1
    assert "introuvable" in capsys.readouterr().err.lower()
