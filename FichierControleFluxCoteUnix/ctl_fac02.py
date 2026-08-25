#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ctl_fac02.py - Controle du flux FAC02 (factures clients) - version Windows/Python

Usage : python ctl_fac02.py <fichier_SRC> [fichier_CTL]
        Si le fichier CTL est omis, il est deduit du nom du fichier SRC
        (remplacement de _SRC_ par _CTL_).

Traitements :
  1. Synthese par portefeuille : nb de factures (lignes comptes 411*)
     et montant total (Debit +, Credit -, montants en centimes / 100)
  2. Rapprochement SRC <-> CTL (nb factures + montant par portefeuille)
  3. Liste des factures a montant 0
  4. Generation d'un rapport Excel FAC02_SYNTHESE_<horodatage>.xlsx
     dans le repertoire du fichier SRC (si openpyxl est installe)

Code retour : 0 = rapprochement OK, 1 = ecart detecte, 2 = erreur usage
"""

import os
import sys
from collections import defaultdict

from rapport_excel import chemin_rapport


def fr(montant):
    """Format francais : virgule decimale."""
    return ("%.2f" % montant).replace(".", ",")


def lire_src(chemin):
    """Agrege les lignes comptes clients 411* par portefeuille."""
    cnt = defaultdict(int)
    amt = defaultdict(int)          # en centimes pour rester exact
    zeros = []
    with open(chemin, "r", encoding="latin-1") as f:
        for ligne in f:
            ch = ligne.rstrip("\r\n").split(";")
            if len(ch) < 17 or not ch[6].startswith("411"):
                continue
            portefeuille, facture, compte, date = ch[16], ch[2], ch[6], ch[8]
            m = int(ch[12] or 0)
            if ch[13] == "-":
                m = -m
            if ch[11] == "C":
                m = -m
            cnt[portefeuille] += 1
            amt[portefeuille] += m
            if m == 0:
                zeros.append((portefeuille, facture, compte, date))
    return cnt, amt, zeros


def lire_ctl(chemin):
    """CTL : PORTEFEUILLE;DATE;NB;MONTANT;MONTANT;FICHIER;0"""
    attendu = {}
    with open(chemin, "r", encoding="latin-1") as f:
        for ligne in f:
            ch = ligne.rstrip("\r\n").split(";")
            if len(ch) >= 4 and ch[0]:
                attendu[ch[0]] = (int(ch[2]), float(ch[3].replace(",", ".")))
    return attendu


def generer_excel(src, ctl, rappro, zeros, ecart,
                  titre="CONTROLE FLUX FAC02 - FACTURES CLIENTS",
                  libelle="Portefeuille", prefixe="FAC02_SYNTHESE"):
    """Rapport Excel : onglet synthese/rapprochement + onglet factures a 0."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        print("INFO : openpyxl non installe, rapport Excel non genere "
              "(pip install openpyxl)", file=sys.stderr)
        return None

    chemin = chemin_rapport(src, prefixe)
    os.makedirs(os.path.dirname(chemin), exist_ok=True)

    bleu = "1F497D"
    bleu_clair = "EAF2F8"
    peche = "FCE4D6"
    vert = "EDEFDA"
    gris = "666666"
    gras = Font(bold=True)
    entete_fill = PatternFill("solid", fgColor=bleu)
    entete_font = Font(bold=True, color="FFFFFF")
    ok_fill = PatternFill("solid", fgColor=vert)
    ko_fill = PatternFill("solid", fgColor=peche)
    bande_fill = PatternFill("solid", fgColor=bleu_clair)
    centre = Alignment(horizontal="center", vertical="center")
    bordure = Border(bottom=Side(style="thin", color="D9D9D9"))

    wb = Workbook()

    ws = wb.active
    ws.title = "Synthese"
    ws.append([titre])
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = entete_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.append(["Fichier SRC", os.path.basename(src)])
    ws.append(["Fichier CTL", os.path.basename(ctl)])
    ws.append(["Resultat", "ECART(S) DETECTE(S)" if ecart else "RAPPROCHEMENT OK"])
    for lig in (2, 3, 4):
        ws.cell(row=lig, column=1).font = gras
        ws.cell(row=lig, column=1).fill = bande_fill
    ws["B2"].font = Font(italic=True, color=gris)
    ws["B3"].font = Font(italic=True, color=gris)
    ws.cell(row=4, column=2).fill = ko_fill if ecart else ok_fill
    ws.cell(row=4, column=2).font = gras
    ws.append([])

    entetes = [libelle, "Nb factures SRC", "Montant SRC",
               "Nb factures CTL", "Montant CTL", "Statut"]
    ws.append(entetes)
    lig_entete = ws.max_row
    for col in range(1, len(entetes) + 1):
        c = ws.cell(row=lig_entete, column=col)
        c.font, c.fill, c.alignment = entete_font, entete_fill, centre
        c.border = bordure
    ws.row_dimensions[lig_entete].height = 32

    tot_cnt = tot_amt = 0
    for r in rappro:
        ws.append([r["ptf"], r["src_cnt"], r["src_amt"],
                   r["ctl_cnt"], r["ctl_amt"], r["statut"]])
        c = ws.cell(row=ws.max_row, column=6)
        c.fill = ok_fill if r["statut"] == "OK" else ko_fill
        c.alignment = centre
        c.font = gras
        if (ws.max_row - lig_entete) % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=ws.max_row, column=col).fill = bande_fill
        for col in range(1, 7):
            ws.cell(row=ws.max_row, column=col).border = bordure
        tot_cnt += r["src_cnt"] or 0
        tot_amt += r["src_amt"] or 0
    ws.append(["TOTAL", tot_cnt, round(tot_amt, 2), None, None, None])
    for col in range(1, 7):
        ws.cell(row=ws.max_row, column=col).font = gras
        ws.cell(row=ws.max_row, column=col).fill = bande_fill
        ws.cell(row=ws.max_row, column=col).border = bordure
    for lig in range(lig_entete + 1, ws.max_row + 1):
        for col in (3, 5):
            ws.cell(row=lig, column=col).number_format = '#,##0.00;[Red](#,##0.00);-'
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = "A%d:F%d" % (lig_entete, ws.max_row - 1)
    ws.sheet_view.showGridLines = False
    for col, largeur in zip("ABCDEF", (18, 19, 18, 19, 18, 14)):
        ws.column_dimensions[col].width = largeur

    ws2 = wb.create_sheet("Factures a 0")
    ws2.append(["FACTURES A MONTANT ZERO"])
    ws2.merge_cells("A1:D1")
    ws2["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws2["A1"].fill = entete_fill
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 28
    ws2.append(["Fichier source", os.path.basename(src)])
    ws2["A2"].font = gras
    ws2["A2"].fill = bande_fill
    ws2["B2"].font = Font(italic=True, color=gris)
    ws2.append([])
    entetes2 = [libelle, "N° facture", "Compte", "Date piece"]
    ws2.append(entetes2)
    for col in range(1, len(entetes2) + 1):
        c = ws2.cell(row=4, column=col)
        c.font, c.fill, c.alignment = entete_font, entete_fill, centre
        c.border = bordure
    ws2.row_dimensions[4].height = 30
    if zeros:
        for index, (p, facture, compte, date) in enumerate(zeros, 1):
            ws2.append([p, facture, compte, date])
            if index % 2 == 0:
                for col in range(1, 5):
                    ws2.cell(row=ws2.max_row, column=col).fill = bande_fill
            for col in range(1, 5):
                ws2.cell(row=ws2.max_row, column=col).border = bordure
    else:
        ws2.append(["Aucune facture a montant 0", None, None, None])
        ws2["A5"].fill = ok_fill
        ws2["A5"].font = gras
    ws2.freeze_panes = "A5"
    ws2.auto_filter.ref = "A4:D%d" % ws2.max_row
    ws2.sheet_view.showGridLines = False
    for col, largeur in zip("ABCD", (18, 20, 16, 16)):
        ws2.column_dimensions[col].width = largeur

    try:
        wb.save(chemin)
    except PermissionError:
        print("ATTENTION : impossible d'ecrire %s (fichier ouvert dans "
              "Excel ?). Rapport non genere." % chemin, file=sys.stderr)
        return None
    return chemin


def main():
    if len(sys.argv) < 2:
        print("Usage : %s <fichier_SRC> [fichier_CTL]" % sys.argv[0], file=sys.stderr)
        return 2
    src = sys.argv[1]
    ctl = sys.argv[2] if len(sys.argv) > 2 else src.replace("_SRC_", "_CTL_")
    for chemin in (src, ctl):
        if not os.path.isfile(chemin):
            print("ERREUR : fichier introuvable : %s" % chemin, file=sys.stderr)
            return 2

    print("=" * 77)
    print(" CONTROLE FLUX FAC02 - FACTURES CLIENTS")
    print(" Fichier SRC : %s" % src)
    print(" Fichier CTL : %s" % ctl)
    print("=" * 77)

    cnt, amt, zeros = lire_src(src)
    attendu = lire_ctl(ctl)

    print("\n--- 1. SYNTHESE PAR PORTEFEUILLE (fichier SRC) ---")
    for p in sorted(cnt):
        print("Portefeuille %-4s : %5d factures, montant %15s"
              % (p, cnt[p], fr(amt[p] / 100.0)))

    print("\n--- 2. RAPPROCHEMENT SRC <-> CTL ---")
    ecart = False
    rappro = []
    for p in sorted(set(cnt) | set(attendu)):
        ligne = {"ptf": p, "src_cnt": cnt.get(p), "src_amt": None,
                 "ctl_cnt": None, "ctl_amt": None, "statut": "ECART"}
        if p in cnt:
            ligne["src_amt"] = round(amt[p] / 100.0, 2)
        if p in attendu:
            ligne["ctl_cnt"], ligne["ctl_amt"] = attendu[p]
        if p not in attendu:
            print("ECART  %-4s : present dans SRC mais absent du CTL (%d factures)"
                  % (p, cnt[p]))
            ecart = True
        elif p not in cnt:
            print("ECART  %-4s : present dans CTL mais absent du SRC" % p)
            ligne["src_cnt"] = 0
            ecart = True
        else:
            sa = amt[p] / 100.0
            nb_ctl, amt_ctl = attendu[p]
            if cnt[p] == nb_ctl and abs(sa - amt_ctl) < 0.005:
                print("OK     %-4s : %5d factures, montant %s" % (p, cnt[p], fr(sa)))
                ligne["statut"] = "OK"
            else:
                print("ECART  %-4s : SRC=%d factures / %s  --  CTL=%d factures / %s"
                      % (p, cnt[p], fr(sa), nb_ctl, fr(amt_ctl)))
                ecart = True
        rappro.append(ligne)

    print("\n--- 3. FACTURES A MONTANT 0 ---")
    if not zeros:
        print("Aucune facture a montant 0.")
    else:
        for p, facture, compte, date in zeros:
            print("Portefeuille %-4s facture %-12s compte %s date %s"
                  % (p, facture, compte, date))
        print("TOTAL : %d facture(s) a montant 0" % len(zeros))

    print("\n--- 4. RAPPORT EXCEL ---")
    xlsx = generer_excel(src, ctl, rappro, zeros, ecart)
    if xlsx:
        print("Rapport genere : %s" % xlsx)

    print("\n" + "=" * 77)
    if ecart:
        print(" RESULTAT : ECART(S) DETECTE(S) - voir section 2")
        print("=" * 77)
        return 1
    print(" RESULTAT : RAPPROCHEMENT OK")
    print("=" * 77)
    return 0


if __name__ == "__main__":
    sys.exit(main())
