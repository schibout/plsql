#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Controle local du flux d'ecritures GL : rapprochement SRC / CTL."""

import csv
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Set, Tuple

from rapport_excel import chemin_rapport, nom_onglet
from selectionner_source import ErreurSelection, selectionner_source


EXIT_OK = 0
EXIT_TECHNIQUE = 1
EXIT_ANOMALIE = 2


class ErreurControle(Exception):
    """Erreur technique ou fichier d'entree invalide."""


@dataclass
class AgregatSrc:
    pieces: Set[str] = field(default_factory=set)
    debit: int = 0
    credit: int = 0
    lignes: int = 0


@dataclass
class AttenduCtl:
    nb: int
    debit: int
    credit: int
    nom_src: str


@dataclass
class ResultatOrigine:
    origine: str
    nb_src: Optional[int]
    debit_src: Optional[int]
    credit_src: Optional[int]
    nb_ctl: Optional[int]
    debit_ctl: Optional[int]
    credit_ctl: Optional[int]
    statut: str


@dataclass
class ResultatControle:
    src: Path
    ctl: Path
    origines: Dict[str, ResultatOrigine]
    pieces_desequilibrees: List[Tuple[str, str, int, int]]
    code_retour: int
    rapport: Optional[Path] = None


def centimes(texte: str, fichier: Path, numero_ligne: int) -> int:
    valeur = texte.strip().replace(" ", "")
    if not valeur:
        return 0
    try:
        montant = Decimal(valeur.replace(",", "."))
    except InvalidOperation as exc:
        raise ErreurControle(
            "montant illisible ligne %d dans %s : '%s'"
            % (numero_ligne, fichier, texte)
        ) from exc
    en_centimes = montant * 100
    if en_centimes != en_centimes.to_integral_value():
        raise ErreurControle(
            "montant avec plus de deux decimales ligne %d dans %s : '%s'"
            % (numero_ligne, fichier, texte)
        )
    return int(en_centimes)


def format_montant(valeur: Optional[int]) -> str:
    if valeur is None:
        return "-"
    signe = "-" if valeur < 0 else ""
    absolu = abs(valeur)
    euros, cents = divmod(absolu, 100)
    return "%s%s,%02d" % (signe, format(euros, ",").replace(",", " "), cents)


def selectionner_src(entree=None) -> Path:
    dossier_script = Path(__file__).resolve().parent
    racine = entree if entree is not None and str(entree).strip() else dossier_script
    try:
        return selectionner_source("GL", racine)
    except ErreurSelection as exc:
        raise ErreurControle(str(exc)) from exc


def deduire_ctl(src: Path) -> Path:
    texte = str(src)
    if "_SRC_" not in texte:
        raise ErreurControle("impossible de deduire le CTL du nom SRC : %s" % src)
    return Path(texte.replace("_SRC_", "_CTL_"))


def lire_src(src: Path):
    aggregats: Dict[str, AgregatSrc] = defaultdict(AgregatSrc)
    pieces = defaultdict(lambda: [0, 0])

    try:
        fichier = src.open("r", encoding="latin-1", newline="")
    except OSError as exc:
        raise ErreurControle("impossible de lire %s : %s" % (src, exc)) from exc

    with fichier:
        for numero, champs in enumerate(csv.reader(fichier, delimiter=";"), 1):
            if not champs or all(not champ.strip() for champ in champs):
                continue
            if len(champs) < 12:
                raise ErreurControle(
                    "ligne %d dans %s : moins de 12 colonnes" % (numero, src)
                )
            piece = champs[0].strip()
            origine = champs[11].strip()
            if not piece:
                raise ErreurControle("numero de piece vide ligne %d dans %s" % (numero, src))
            if not origine:
                raise ErreurControle("origine vide ligne %d dans %s" % (numero, src))
            debit = centimes(champs[8], src, numero)
            credit = centimes(champs[9], src, numero)

            agregat = aggregats[origine]
            agregat.pieces.add(piece)
            agregat.debit += debit
            agregat.credit += credit
            agregat.lignes += 1
            pieces[(origine, piece)][0] += debit
            pieces[(origine, piece)][1] += credit

    if not aggregats:
        raise ErreurControle("aucune ecriture GL exploitable dans %s" % src)

    desequilibrees = sorted(
        (origine, piece, montants[0], montants[1])
        for (origine, piece), montants in pieces.items()
        if montants[0] != montants[1]
    )
    return dict(aggregats), desequilibrees


def lire_ctl(ctl: Path) -> Dict[str, AttenduCtl]:
    attendus = {}
    try:
        fichier = ctl.open("r", encoding="latin-1", newline="")
    except OSError as exc:
        raise ErreurControle("impossible de lire %s : %s" % (ctl, exc)) from exc

    with fichier:
        for numero, champs in enumerate(csv.reader(fichier, delimiter=";"), 1):
            if not champs or all(not champ.strip() for champ in champs):
                continue
            if len(champs) < 6:
                raise ErreurControle(
                    "ligne %d dans %s : moins de 6 colonnes" % (numero, ctl)
                )
            origine = champs[0].strip()
            if not origine:
                raise ErreurControle("origine vide ligne %d dans %s" % (numero, ctl))
            if origine in attendus:
                raise ErreurControle("origine dupliquee dans le CTL ligne %d : %s" % (numero, origine))
            try:
                nb = int(champs[2].strip())
            except ValueError as exc:
                raise ErreurControle(
                    "nombre de pieces illisible ligne %d dans %s : '%s'"
                    % (numero, ctl, champs[2])
                ) from exc
            attendus[origine] = AttenduCtl(
                nb=nb,
                debit=centimes(champs[3], ctl, numero),
                credit=centimes(champs[4], ctl, numero),
                nom_src=champs[5].strip(),
            )

    if not attendus:
        raise ErreurControle("aucune attente exploitable dans %s" % ctl)
    return attendus


def rapprocher(src: Path, ctl: Path) -> ResultatControle:
    aggregats, desequilibrees = lire_src(src)
    attendus = lire_ctl(ctl)
    origines = {}
    origines_desequilibrees = {origine for origine, _, _, _ in desequilibrees}

    for origine in sorted(set(aggregats) | set(attendus)):
        agregat = aggregats.get(origine)
        attendu = attendus.get(origine)
        conforme = bool(agregat and attendu)
        if conforme:
            conforme = (
                len(agregat.pieces) == attendu.nb
                and agregat.debit == attendu.debit
                and agregat.credit == attendu.credit
                and agregat.debit == agregat.credit
                and origine not in origines_desequilibrees
            )
        origines[origine] = ResultatOrigine(
            origine=origine,
            nb_src=len(agregat.pieces) if agregat else None,
            debit_src=agregat.debit if agregat else None,
            credit_src=agregat.credit if agregat else None,
            nb_ctl=attendu.nb if attendu else None,
            debit_ctl=attendu.debit if attendu else None,
            credit_ctl=attendu.credit if attendu else None,
            statut="OK" if conforme else "ECART",
        )

    anomalie = bool(desequilibrees) or any(
        ligne.statut != "OK" for ligne in origines.values()
    )
    return ResultatControle(
        src=src,
        ctl=ctl,
        origines=origines,
        pieces_desequilibrees=desequilibrees,
        code_retour=EXIT_ANOMALIE if anomalie else EXIT_OK,
    )


def generer_excel(resultat: ResultatControle) -> Optional[Path]:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        print("INFO : openpyxl non installe, rapport Excel non genere", file=sys.stderr)
        return None

    chemin = Path(chemin_rapport(resultat.src, "GL_SYNTHESE"))
    chemin.parent.mkdir(parents=True, exist_ok=True)

    # Le classeur du flux est partage entre les exports : chaque export y ecrit
    # ses propres onglets (suffixes par son horodatage) sans toucher aux autres.
    nom_synthese = nom_onglet("Synthese", resultat.src)
    nom_desequilibrees = nom_onglet("Pieces desequilibrees", resultat.src)

    bleu = "1F497D"
    bleu_clair = "EAF2F8"
    peche = "FCE4D6"
    vert = "EDEFDA"
    entete_fill = PatternFill("solid", fgColor=bleu)
    entete_font = Font(bold=True, color="FFFFFF")
    ok_fill = PatternFill("solid", fgColor=vert)
    ecart_fill = PatternFill("solid", fgColor=peche)
    bande_fill = PatternFill("solid", fgColor=bleu_clair)
    centre = Alignment(horizontal="center", vertical="center")
    bordure = Border(bottom=Side(style="thin", color="D9D9D9"))

    if chemin.is_file():
        try:
            wb = load_workbook(chemin)
        except Exception:
            wb = Workbook()
            wb.remove(wb.active)
    else:
        wb = Workbook()
        wb.remove(wb.active)
    for nom in (nom_synthese, nom_desequilibrees):
        if nom in wb.sheetnames:
            del wb[nom]

    ws = wb.create_sheet(nom_synthese)
    ws.append(["CONTROLE ECRITURES GL"])
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = entete_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.append(["Fichier SRC", resultat.src.name])
    ws.append(["Fichier CTL", resultat.ctl.name])
    ws.append(["Resultat", "OK" if resultat.code_retour == 0 else "ANOMALIE(S)"])
    for ligne_meta in (2, 3, 4):
        ws.cell(ligne_meta, 1).font = Font(bold=True)
        ws.cell(ligne_meta, 1).fill = bande_fill
    ws.cell(4, 2).fill = ok_fill if resultat.code_retour == 0 else ecart_fill
    ws.cell(4, 2).font = Font(bold=True)
    ws.append([])
    entetes = [
        "Origine", "Nb pieces SRC", "Debit SRC", "Credit SRC",
        "Nb pieces CTL", "Debit CTL", "Credit CTL", "Statut",
    ]
    ws.append(entetes)
    for cellule in ws[ws.max_row]:
        cellule.font = entete_font
        cellule.fill = entete_fill
        cellule.alignment = centre
        cellule.border = bordure
    ws.row_dimensions[ws.max_row].height = 32
    for index, ligne in enumerate(resultat.origines.values(), 1):
        ws.append([
            ligne.origine,
            ligne.nb_src,
            None if ligne.debit_src is None else ligne.debit_src / 100,
            None if ligne.credit_src is None else ligne.credit_src / 100,
            ligne.nb_ctl,
            None if ligne.debit_ctl is None else ligne.debit_ctl / 100,
            None if ligne.credit_ctl is None else ligne.credit_ctl / 100,
            ligne.statut,
        ])
        if index % 2 == 0:
            for colonne in range(1, 8):
                ws.cell(ws.max_row, colonne).fill = bande_fill
        for colonne in range(1, 9):
            ws.cell(ws.max_row, colonne).border = bordure
        ws.cell(ws.max_row, 8).fill = ok_fill if ligne.statut == "OK" else ecart_fill
        ws.cell(ws.max_row, 8).font = Font(bold=True)
        ws.cell(ws.max_row, 8).alignment = centre
    for colonne in (3, 4, 6, 7):
        for ligne in range(7, ws.max_row + 1):
            ws.cell(ligne, colonne).number_format = '#,##0.00;[Red](#,##0.00);-'
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = "A6:H%d" % ws.max_row
    ws.sheet_view.showGridLines = False
    for colonne, largeur in zip("ABCDEFGH", (16, 18, 18, 18, 18, 18, 18, 14)):
        ws.column_dimensions[colonne].width = largeur

    wd = wb.create_sheet(nom_desequilibrees)
    wd.append(["PIECES GL DESEQUILIBREES"])
    wd.merge_cells("A1:E1")
    wd["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    wd["A1"].fill = entete_fill
    wd["A1"].alignment = Alignment(horizontal="left", vertical="center")
    wd.row_dimensions[1].height = 28
    wd.append(["Fichier source", resultat.src.name])
    wd["A2"].font = Font(bold=True)
    wd["A2"].fill = bande_fill
    wd.append([])
    wd.append(["Origine", "Numero piece", "Debit", "Credit", "Ecart"])
    for cellule in wd[4]:
        cellule.font = entete_font
        cellule.fill = entete_fill
        cellule.alignment = centre
        cellule.border = bordure
    for index, (origine, piece, debit, credit) in enumerate(resultat.pieces_desequilibrees, 1):
        wd.append([origine, piece, debit / 100, credit / 100, (debit - credit) / 100])
        if index % 2 == 0:
            for colonne in range(1, 5):
                wd.cell(wd.max_row, colonne).fill = bande_fill
        wd.cell(wd.max_row, 5).fill = ecart_fill
        for colonne in range(1, 6):
            wd.cell(wd.max_row, colonne).border = bordure
    if not resultat.pieces_desequilibrees:
        wd.append(["Aucune piece desequilibree", None, None, None, None])
        wd["A5"].fill = ok_fill
        wd["A5"].font = Font(bold=True)
    for ligne in range(5, wd.max_row + 1):
        for colonne in (3, 4, 5):
            wd.cell(ligne, colonne).number_format = '#,##0.00;[Red](#,##0.00);-'
    wd.freeze_panes = "A5"
    wd.auto_filter.ref = "A4:E%d" % wd.max_row
    wd.sheet_view.showGridLines = False
    for colonne, largeur in zip("ABCDE", (16, 20, 18, 18, 18)):
        wd.column_dimensions[colonne].width = largeur

    try:
        wb.save(chemin)
    except (OSError, PermissionError) as exc:
        print("ATTENTION : rapport Excel non genere : %s" % exc, file=sys.stderr)
        return None
    return chemin


def afficher(resultat: ResultatControle) -> None:
    print("=" * 101)
    print(" CONTROLE FLUX - ECRITURES GL")
    print(" Fichier SRC : %s" % resultat.src)
    print(" Fichier CTL : %s" % resultat.ctl)
    print("=" * 101)
    print(
        "%-8s | %7s | %14s | %14s | %7s | %14s | %14s | %s"
        % ("ORIGINE", "NB SRC", "DEBIT SRC", "CREDIT SRC", "NB CTL", "DEBIT CTL", "CREDIT CTL", "STATUT")
    )
    print("-" * 101)
    for ligne in resultat.origines.values():
        print(
            "%-8s | %7s | %14s | %14s | %7s | %14s | %14s | %s"
            % (
                ligne.origine,
                "-" if ligne.nb_src is None else ligne.nb_src,
                format_montant(ligne.debit_src),
                format_montant(ligne.credit_src),
                "-" if ligne.nb_ctl is None else ligne.nb_ctl,
                format_montant(ligne.debit_ctl),
                format_montant(ligne.credit_ctl),
                ligne.statut,
            )
        )

    total_debit = sum(ligne.debit_src or 0 for ligne in resultat.origines.values())
    total_credit = sum(ligne.credit_src or 0 for ligne in resultat.origines.values())
    print("-" * 101)
    print(" TOTAL SRC : debit %s / credit %s" % (format_montant(total_debit), format_montant(total_credit)))
    if resultat.pieces_desequilibrees:
        print("\n PIECES DESEQUILIBREES :")
        for origine, piece, debit, credit in resultat.pieces_desequilibrees:
            print(
                "   %s / %s : debit %s, credit %s"
                % (origine, piece, format_montant(debit), format_montant(credit))
            )
    if resultat.rapport:
        print(" Rapport Excel : %s" % resultat.rapport)
    print("=" * 101)
    print(" RESULTAT : %s" % ("CONTROLE OK" if resultat.code_retour == 0 else "ANOMALIE(S) DETECTEE(S)"))


def controler(src, ctl=None, generer_rapport=True) -> ResultatControle:
    chemin_src = Path(src).resolve()
    chemin_ctl = Path(ctl).resolve() if ctl is not None else deduire_ctl(chemin_src)
    if not chemin_src.is_file():
        raise ErreurControle("fichier SRC introuvable : %s" % chemin_src)
    if not chemin_ctl.is_file():
        raise ErreurControle("fichier CTL introuvable : %s" % chemin_ctl)
    resultat = rapprocher(chemin_src, chemin_ctl)
    if generer_rapport:
        resultat.rapport = generer_excel(resultat)
    return resultat


def executer(arguments: Optional[Sequence[str]] = None, generer_rapport=True) -> int:
    args = list(sys.argv[1:] if arguments is None else arguments)
    if len(args) > 2:
        print("ERREUR : usage : ctl_ecritures_gl.py [dossier|fichier_SRC] [fichier_CTL]", file=sys.stderr)
        return EXIT_TECHNIQUE
    try:
        src = selectionner_src(args[0] if args else None)
        ctl = Path(args[1]).expanduser().resolve() if len(args) == 2 else deduire_ctl(src)
        resultat = controler(src, ctl, generer_rapport=generer_rapport)
        afficher(resultat)
        return resultat.code_retour
    except ErreurControle as exc:
        print("ERREUR : %s" % exc, file=sys.stderr)
        return EXIT_TECHNIQUE


if __name__ == "__main__":
    sys.exit(executer())
