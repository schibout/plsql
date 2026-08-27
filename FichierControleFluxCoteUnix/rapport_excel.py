#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
rapport_excel.py - Onglets ajoutes au classeur de synthese d'un flux

Ce module regroupe ce qui est commun aux rapports Excel des controles :
  - la palette et la mise en forme des onglets ;
  - la regle de nommage du classeur de synthese (chemin_rapport) ;
  - l'ajout d'onglets a un classeur existant (ajouter_onglets).

Il sert aussi de passerelle pour les scripts PowerShell Verifier_Oracle_*.ps1 :
ceux-ci ecrivent le resultat de leurs interrogations Oracle dans un fichier
JSON, puis appellent

    python rapport_excel.py <fichier_json>

qui ajoute les onglets decrits au classeur de synthese produit juste avant par
le controle local. Le rapport du flux reste ainsi un fichier unique, et aucune
installation d'Excel n'est necessaire (openpyxl suffit).

Structure attendue du JSON :

    {
      "src": "...\\SOURCE\\..._SRC_....csv",
      "prefixe": "FAC02_SYNTHESE_FOURNISSEURS",
      "onglets": [
        {
          "nom": "Detail Oracle",
          "titre": "CONTROLE FAC02 FOURNISSEURS - DETAIL ORACLE",
          "sous_titre": "Execution : 25/08/2026 11:00:00",
          "colonnes": ["Folio", "Numero Piece", "Montant Fichier", "Statut"],
          "colonne_statut": "Statut",
          "valeurs_ok": ["INTEGREE"],
          "lignes": [{"Folio": "CEC", "Numero Piece": "12345", ...}]
        }
      ]
    }

Code retour : 0 = onglets ajoutes, 1 = erreur technique
"""

import json
import os
import re
import sys


BLEU = "1F497D"
BLEU_CLAIR = "EAF2F8"
PECHE = "FCE4D6"
VERT = "EDEFDA"
GRIS = "666666"

# Colonnes reconnues par leur intitule, pour le format des nombres.
FORMAT_MONTANT = '#,##0.00;[Red](#,##0.00);-'
FORMAT_ENTIER = '#,##0;[Red](#,##0);-'
PREFIXES_MONTANT = ("montant", "debit", "credit", "ecart")
PREFIXES_ENTIER = ("nb ",)


def nom_export(src):
    """Nom du dossier de flux d'un export, ou None hors de cette arborescence.

    C'est le dossier de flux, celui passe aux lanceurs controle*.bat, qui
    nomme le rapport. Le nom est deduit du chemin du SRC et non de l'argument
    saisi, pour rester le meme que l'on passe le dossier, le fichier, ou rien.

    Deux dispositions coexistent selon l'anciennete de l'export :

        <flux>\\<instance>\\SOURCE\\<SRC>
        <flux>\\SOURCE|TARGET\\<instance>\\SOURCE\\<SRC>   (deux demi-flux)

    La seconde est celle que rapatrient les copier_instances_*.sh. Sans le
    saut du niveau de demi-flux, tous les exports se nommeraient SOURCE et
    partageraient un unique classeur, en s'ecrasant les uns les autres.
    """
    chemin = os.path.abspath(str(src))
    source = os.path.dirname(chemin)
    if os.path.basename(source).upper() != "SOURCE":
        return None
    flux = os.path.dirname(os.path.dirname(source))
    if os.path.basename(flux).upper() in ("SOURCE", "TARGET"):
        flux = os.path.dirname(flux)
    nom = os.path.basename(flux)
    # os.path.basename d'une racine de disque ("C:\\") est vide.
    return nom or None


def dossier_rapport():
    """Repertoire des classeurs de synthese."""
    return os.environ.get("CONTROLE_FLUX_RAPPORT_DIR") or os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "rapport"
    )


def chemin_rapport(src, prefixe):
    """Chemin du classeur de synthese d'un flux, deduit du chemin du SRC.

    La regle doit rester deterministe : le controle local cree le classeur,
    puis le controle Oracle le retrouve pour y ajouter ses onglets.

    Le classeur porte le seul nom du dossier de flux, celui passe a
    controle.bat : 10082026_FOURNISSEUR_CEG.xlsx. Ce nom identifie deja la
    date, le type de flux et le code, le prefixe n'apporterait rien.
    Pour un SRC isole hors d'un dossier d'export, le prefixe du flux et
    l'horodatage du fichier prennent le relais.
    """
    dossier = dossier_rapport()
    nom = nom_export(src)
    if not nom:
        correspondance = re.search(
            r"_(\d{6}-\d{6})(?:_|\.)", os.path.basename(str(src))
        )
        suffixe = correspondance.group(1) if correspondance else "rapport"
        nom = "%s_%s" % (prefixe, suffixe)
    return os.path.join(dossier, "%s.xlsx" % nom)


def suffixe_export(src):
    """Horodatage extrait du nom du SRC (ex. "170826_161552"), ou None.

    Plusieurs exports d'un meme dossier de flux partagent le meme classeur :
    cet horodatage suffixe les noms d'onglets pour que chaque export garde
    les siens au lieu d'ecraser ceux du precedent.
    """
    correspondance = re.search(r"_(\d{6}[-_]\d{6})(?=[_.])", os.path.basename(str(src)))
    return correspondance.group(1) if correspondance else None


def nom_onglet(base, src):
    """Nom d'onglet suffixe par l'horodatage de l'export, limite a 31 car."""
    suffixe = suffixe_export(src)
    if not suffixe:
        return base[:31]
    return "%s %s" % (base[: 31 - len(suffixe) - 1], suffixe)


def format_colonne(intitule):
    """Format Excel deduit de l'intitule de colonne (None = format general)."""
    minuscule = str(intitule).strip().lower()
    if minuscule.startswith(PREFIXES_MONTANT):
        return FORMAT_MONTANT
    if minuscule.startswith(PREFIXES_ENTIER):
        return FORMAT_ENTIER
    return None


def ecrire_onglet(ws, titre, sous_titre, colonnes, lignes,
                  colonne_statut=None, valeurs_ok=()):
    """Remplit une feuille : titre, sous-titre, en-tete fige et lignes zebrees.

    Meme disposition que les onglets produits par les controles locaux :
    ligne 1 titre, ligne 2 sous-titre, ligne 4 en-tete, donnees a partir de 5.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    entete_fill = PatternFill("solid", fgColor=BLEU)
    entete_font = Font(bold=True, color="FFFFFF")
    ok_fill = PatternFill("solid", fgColor=VERT)
    ko_fill = PatternFill("solid", fgColor=PECHE)
    bande_fill = PatternFill("solid", fgColor=BLEU_CLAIR)
    centre = Alignment(horizontal="center", vertical="center")
    bordure = Border(bottom=Side(style="thin", color="D9D9D9"))
    gras = Font(bold=True)

    nb_colonnes = max(len(colonnes), 1)
    ws.cell(1, 1, titre)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_colonnes)
    ws.cell(1, 1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(1, 1).fill = entete_fill
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    for colonne in range(1, nb_colonnes + 1):
        ws.cell(1, colonne).fill = entete_fill

    ws.cell(2, 1, sous_titre)
    ws.cell(2, 1).font = Font(italic=True, color=GRIS)

    for index, intitule in enumerate(colonnes, 1):
        cellule = ws.cell(4, index, intitule)
        cellule.font = entete_font
        cellule.fill = entete_fill
        cellule.alignment = centre
        cellule.border = bordure
    ws.row_dimensions[4].height = 34

    index_statut = None
    if colonne_statut and colonne_statut in colonnes:
        index_statut = colonnes.index(colonne_statut) + 1
    valeurs_ok = set(valeurs_ok or ())

    for rang, ligne in enumerate(lignes, 1):
        numero = 4 + rang
        for index, intitule in enumerate(colonnes, 1):
            valeur = ligne.get(intitule) if isinstance(ligne, dict) else None
            # Une valeur absente cote PowerShell peut arriver en objet vide.
            if isinstance(valeur, (dict, list, tuple)):
                valeur = None
            cellule = ws.cell(numero, index, valeur)
            cellule.border = bordure
            forme = format_colonne(intitule)
            if forme:
                cellule.number_format = forme
            if rang % 2 == 0:
                cellule.fill = bande_fill
        if index_statut:
            cellule = ws.cell(numero, index_statut)
            cellule.fill = ok_fill if cellule.value in valeurs_ok else ko_fill
            cellule.font = gras
            cellule.alignment = centre

    if not lignes:
        cellule = ws.cell(5, 1, "Aucune donnee")
        cellule.fill = ok_fill
        cellule.font = gras

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:%s%d" % (
        get_column_letter(nb_colonnes), max(ws.max_row, 5)
    )
    ws.sheet_view.showGridLines = False
    for index, intitule in enumerate(colonnes, 1):
        largeur = min(28, max(12, len(str(intitule)) + 4))
        ws.column_dimensions[get_column_letter(index)].width = largeur
    return ws


def ajouter_onglets(classeur, onglets, creer=False):
    """Ajoute (ou remplace) des onglets dans un classeur.

    Le classeur doit exister : il est cree par le controle local, qui passe
    toujours avant les controles Oracle. Le drapeau creer leve cette
    contrainte pour la reconciliation, qui peut etre lancee seule sur un
    export dont le controle local n'a pas ete rejoue.
    """
    from openpyxl import Workbook, load_workbook

    if creer and not os.path.isfile(classeur):
        dossier = os.path.dirname(os.path.abspath(classeur))
        if dossier:
            os.makedirs(dossier, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)
    else:
        wb = load_workbook(classeur)
    for onglet in onglets:
        nom = onglet["nom"][:31]
        if nom in wb.sheetnames:
            del wb[nom]
        ws = wb.create_sheet(nom)
        ecrire_onglet(
            ws,
            onglet.get("titre", nom),
            onglet.get("sous_titre", ""),
            list(onglet.get("colonnes", [])),
            list(onglet.get("lignes", [])),
            onglet.get("colonne_statut"),
            onglet.get("valeurs_ok", ()),
        )
    wb.save(classeur)
    return classeur


def resoudre_classeur(donnees):
    """Retrouve le classeur de synthese vise par un descriptif JSON."""
    chemin = donnees.get("classeur") or chemin_rapport(
        donnees.get("src", ""), donnees.get("prefixe", "SYNTHESE")
    )
    if os.path.isfile(chemin):
        return chemin
    raise IOError("classeur de synthese introuvable : %s" % chemin)


def main(arguments=None):
    args = list(sys.argv[1:] if arguments is None else arguments)
    if len(args) != 1:
        print("ERREUR : usage : rapport_excel.py <fichier_json>", file=sys.stderr)
        return 1
    try:
        with open(args[0], "r", encoding="utf-8") as f:
            donnees = json.load(f)
    except (OSError, ValueError) as exc:
        print("ERREUR : descriptif JSON illisible : %s" % exc, file=sys.stderr)
        return 1

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("INFO : openpyxl non installe, onglets Oracle non ajoutes "
              "(pip install openpyxl)", file=sys.stderr)
        return 1

    try:
        # Les onglets Oracle portent l'horodatage de l'export, comme ceux du
        # controle local, pour cohabiter dans le classeur unique du flux.
        onglets = [
            dict(onglet, nom=nom_onglet(onglet["nom"], donnees.get("src", "")))
            for onglet in donnees.get("onglets", [])
        ]
        classeur = resoudre_classeur(donnees)
        ajouter_onglets(classeur, onglets)
    except (OSError, IOError, KeyError) as exc:
        print("ERREUR : onglets Oracle non ajoutes : %s" % exc, file=sys.stderr)
        return 1
    print(classeur)
    return 0


if __name__ == "__main__":
    sys.exit(main())
